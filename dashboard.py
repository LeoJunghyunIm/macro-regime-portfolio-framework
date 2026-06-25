"""
dashboard.py

Template-based dashboard updater.

This module does NOT create or format dashboard charts.
Instead, it preserves the Excel-designed DASHBOARD sheet and only updates
the hidden/source _DASH_DATA sheet.

Workflow:
1. Copy templates/dashboard_template.xlsx to outputs/Macro_Regime_Framework_Output.xlsx.
2. Export engine/audit sheets into the copied workbook.
3. Update _DASH_DATA so the Excel-designed dashboard charts refresh automatically.

Principle:
- DASHBOARD = human-designed presentation layer.
- _DASH_DATA = Python-updated data contract.
"""

from __future__ import annotations

import shutil
from datetime import date, datetime
from math import ceil, floor
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font


TEMPLATE_PATH = Path("templates/dashboard_template.xlsx")
DASH_DATA_SHEET = "_DASH_DATA"

SHEET_RISK_APPETITE = "Risk Appetite"
SHEET_RATES_INFLATION = "Rates & Inflation"
SHEET_PORTFOLIO_ROUTER = "Portfolio Router"
SHEET_DATA_QUALITY = "Data Quality"
SHEET_STRESS_EVENT_AUDIT = "Stress Event Audit"
SHEET_WEEKLY_DATA = "Weekly Data"


def prepare_output_workbook_from_template(
    output_path: str | Path,
    template_path: str | Path = TEMPLATE_PATH,
) -> bool:
    """
    Copy the Excel dashboard template into the output path.

    Returns True if the template was found and copied.
    Returns False if no template exists, allowing export.py to fall back
    to creating a workbook from scratch.
    """
    output_path = Path(output_path)
    template_path = Path(template_path)

    if not template_path.exists():
        print(
            "Dashboard template not found. "
            "Falling back to standard workbook export."
        )
        return False

    output_path.parent.mkdir(parents=True, exist_ok=True)

    if output_path.exists():
        output_path.unlink()

    shutil.copy2(template_path, output_path)

    print(f"Dashboard template copied to: {output_path}")
    return True


def normalize_date(value):
    """
    Normalize Excel date-like values.
    """
    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    return value


def parse_number(value):
    """
    Convert numeric-like values to float. Return None for blanks / N/A.
    """
    if value is None:
        return None

    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()

    if text.upper() in {"", "N/A", "NA", "NONE", "INSUFFICIENT DATA"}:
        return None

    try:
        return float(text)
    except ValueError:
        return None


def normalize_bool(value) -> bool:
    """
    Convert boolean-like values to bool.
    """
    if value is None:
        return False

    if isinstance(value, bool):
        return value

    if isinstance(value, str):
        return value.strip().lower() in {
            "true",
            "yes",
            "on",
            "1",
            "elevated",
        }

    return bool(value)


def percentile_inc(values: list[float], q: float):
    """
    Excel-like PERCENTILE.INC.

    Used for dashboard visualization bands only.
    Does not change engine logic.
    """
    clean_values = [
        float(value)
        for value in values
        if value is not None
    ]

    if not clean_values:
        return None

    clean_values = sorted(clean_values)

    if len(clean_values) == 1:
        return clean_values[0]

    rank = (len(clean_values) - 1) * q
    lower_index = floor(rank)
    upper_index = ceil(rank)

    if lower_index == upper_index:
        return clean_values[int(rank)]

    lower_value = clean_values[lower_index]
    upper_value = clean_values[upper_index]
    weight = rank - lower_index

    return lower_value + (upper_value - lower_value) * weight


def rolling_percentile(values: list, window: int, q: float) -> list:
    """
    Build rolling percentile series.
    """
    output = []

    for index in range(len(values)):
        if index + 1 < window:
            output.append(None)
            continue

        window_values = values[index + 1 - window:index + 1]
        output.append(percentile_inc(window_values, q))

    return output


def rolling_d4w(values: list) -> list:
    """
    Build 4-week change series.
    """
    output = []

    for index, value in enumerate(values):
        if index < 4:
            output.append(None)
            continue

        prior_value = values[index - 4]

        if value is None or prior_value is None:
            output.append(None)
        else:
            output.append(value - prior_value)

    return output


def latest_marker_values(values: list) -> list:
    """
    Create a helper series where only the latest valid value remains.

    Used for red diamond markers in Excel charts.
    """
    output = [None for _ in values]

    for index in range(len(values) - 1, -1, -1):
        if values[index] is not None:
            output[index] = values[index]
            break

    return output


def read_summary_fields(workbook) -> dict:
    """
    Read SUMMARY sheet into a dictionary.
    """
    if "SUMMARY" not in workbook.sheetnames:
        return {}

    ws = workbook["SUMMARY"]
    summary = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue

        summary[str(row[0])] = row[1] if len(row) > 1 else None

    return summary


def read_sheet_rows(workbook, sheet_name: str) -> list[dict]:
    """
    Read a pandas-exported sheet into list of dictionaries.

    The first column is treated as the date/index column.
    """
    if sheet_name not in workbook.sheetnames:
        return []

    ws = workbook[sheet_name]
    headers = [cell.value for cell in ws[1]]
    rows = []

    for row_number in range(2, ws.max_row + 1):
        first_value = ws.cell(row=row_number, column=1).value

        if first_value is None:
            continue

        row_dict = {
            "Date": normalize_date(first_value),
        }

        for col_number, header in enumerate(headers, start=1):
            if header is None:
                continue

            row_dict[str(header)] = ws.cell(
                row=row_number,
                column=col_number,
            ).value

        rows.append(row_dict)

    return rows


def get_latest_row(rows: list[dict]) -> dict:
    """
    Return the latest row from a list of date-indexed rows.
    """
    if not rows:
        return {}

    return rows[-1]


def risk_appetite_regime_code(regime) -> int | None:
    """
    Risk-On = 0
    Neutral = 1
    Risk-Off = 2
    """
    if regime is None:
        return None

    text = str(regime).strip()

    mapping = {
        "Risk-On": 0,
        "Neutral": 1,
        "Risk-Off": 2,
    }

    return mapping.get(text)


def rates_inflation_regime_code(regime) -> int | None:
    """
    Regime 1 = 1
    Regime 2 = 2
    Regime 3 = 3
    Regime 4 = 4
    """
    if regime is None:
        return None

    if isinstance(regime, (int, float)):
        return int(regime)

    text = str(regime).strip()

    if text.lower().startswith("regime"):
        for part in text.split():
            if part.isdigit():
                return int(part)

    return None


def qa_flag_summary(workbook) -> str:
    """
    Build compact Data Quality status summary from the Data Quality sheet.
    """
    if SHEET_DATA_QUALITY not in workbook.sheetnames:
        return "N/A"

    rows = read_sheet_rows(workbook, SHEET_DATA_QUALITY)

    status_counts = {}

    for row in rows:
        status = row.get("Status")

        if status is None:
            status = row.get("status")

        if status is None:
            continue

        status_counts[str(status)] = status_counts.get(str(status), 0) + 1

    if not status_counts:
        return "N/A"

    parts = [
        f"{status}: {count}"
        for status, count in sorted(status_counts.items())
    ]

    return " | ".join(parts)


def clear_dash_data(ws) -> None:
    """
    Clear the _DASH_DATA data contract area while preserving the sheet itself.

    We intentionally do not delete the sheet because Excel dashboard charts
    reference this sheet.
    """
    max_rows = max(ws.max_row, 1000)
    max_cols = 49  # AW

    for row in ws.iter_rows(
        min_row=1,
        max_row=max_rows,
        min_col=1,
        max_col=max_cols,
    ):
        for cell in row:
            cell.value = None


def is_worksheet_empty(ws) -> bool:
    """
    Return True if a worksheet has no meaningful content.

    Used to remove an empty placeholder DASHBOARD sheet from the output workbook.
    Once the user designs the real dashboard template, the sheet will not be empty
    and will be preserved.
    """
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                return False

    return True


def remove_empty_dashboard_sheet(workbook) -> None:
    """
    Remove DASHBOARD sheet only if it is completely empty.

    This avoids keeping a blank dashboard sheet in the output workbook during
    the template skeleton stage.
    """
    if "DASHBOARD" not in workbook.sheetnames:
        return

    dashboard_ws = workbook["DASHBOARD"]

    if is_worksheet_empty(dashboard_ws):
        workbook.remove(dashboard_ws)
        print("Empty DASHBOARD sheet removed from output workbook.")


def write_table(
    ws,
    start_row: int,
    start_col: int,
    headers: list[str],
    rows: list[list],
) -> None:
    """
    Write a table to _DASH_DATA.
    """
    for col_offset, header in enumerate(headers):
        cell = ws.cell(
            row=start_row,
            column=start_col + col_offset,
        )
        cell.value = header
        cell.font = Font(bold=True)

    for row_offset, row_values in enumerate(rows, start=1):
        for col_offset, value in enumerate(row_values):
            cell = ws.cell(
                row=start_row + row_offset,
                column=start_col + col_offset,
            )
            cell.value = value

            if col_offset == 0 and isinstance(value, (date, datetime)):
                cell.number_format = "m/d/yyyy"


def build_kpi_rows(workbook) -> list[list]:
    """
    Build KPI / Summary block rows for A:B.
    """
    summary = read_summary_fields(workbook)

    risk_appetite_rows = read_sheet_rows(workbook, SHEET_RISK_APPETITE)
    rates_inflation_rows = read_sheet_rows(workbook, SHEET_RATES_INFLATION)

    latest_risk_appetite = get_latest_row(risk_appetite_rows)
    latest_rates_inflation = get_latest_row(rates_inflation_rows)

    rows = [
        ["As of Date", summary.get("As of Date")],
        ["Risk Appetite Regime", summary.get("Risk Appetite Regime")],
        ["Risk Appetite Score", summary.get("Risk Appetite Score")],
        [
            "Risk Appetite Trend",
            latest_risk_appetite.get("Risk Appetite Trend"),
        ],
        [
            "Rates & Inflation Regime",
            summary.get("Rates & Inflation Regime"),
        ],
        ["Rate Pressure", latest_rates_inflation.get("Rate_Pressure")],
        [
            "Inflation Concern",
            latest_rates_inflation.get("Inflation_Concern"),
        ],
        ["Real Yield Tightening", summary.get("Real Yield Tightening")],
        ["Confidence", summary.get("Confidence")],
        ["Equity Stance", summary.get("Equity Stance")],
        ["Equity Band", summary.get("Equity Band")],
        ["Duration Stance", summary.get("Duration Stance")],
        ["Duration Band", summary.get("Duration Band")],
        ["Cash Stance", summary.get("Cash Stance")],
        ["Cash Band", summary.get("Cash Band")],
        ["TIPS Stance", summary.get("TIPS Stance")],
        ["TIPS Band", summary.get("TIPS Band")],
        ["Data Quality Summary", qa_flag_summary(workbook)],
    ]

    return rows


def build_risk_appetite_rows(workbook, max_points: int = 104) -> list[list]:
    """
    Build Risk Appetite chart data block D:I.
    """
    rows = read_sheet_rows(workbook, SHEET_RISK_APPETITE)
    rows = rows[-max_points:]

    scores = [
        parse_number(row.get("Risk Appetite Score"))
        for row in rows
    ]

    latest_scores = latest_marker_values(scores)

    output = []

    for row, score, latest_score in zip(rows, scores, latest_scores):
        regime = row.get("Risk Appetite Regime")

        output.append(
            [
                row.get("Date"),
                score,
                0.5,
                2.5,
                latest_score,
                risk_appetite_regime_code(regime),
            ]
        )

    return output


def build_rates_inflation_timeline_rows(workbook, max_points: int = 104) -> list[list]:
    """
    Build Rates & Inflation regime timeline block K:O.
    """
    rows = read_sheet_rows(workbook, SHEET_RATES_INFLATION)
    rows = rows[-max_points:]

    regime_codes = [
        rates_inflation_regime_code(row.get("Rates & Inflation Regime"))
        for row in rows
    ]

    latest_regime_codes = latest_marker_values(regime_codes)

    output = []

    for row, regime_code, latest_regime_code in zip(
        rows,
        regime_codes,
        latest_regime_codes,
    ):
        output.append(
            [
                row.get("Date"),
                regime_code,
                latest_regime_code,
                1 if normalize_bool(row.get("Rate_Pressure")) else 0,
                1 if normalize_bool(row.get("Inflation_Concern")) else 0,
            ]
        )

    return output


def build_series_rows(
    workbook,
    series_name: str,
    max_points: int = 260,
    rolling_window: int = 52,
) -> list[list]:
    """
    Build DGS10 / T10YIE / DFII10 chart blocks.

    Output columns:
    Date, Series, P80, P20, Latest, 4W Change
    """
    rows = read_sheet_rows(workbook, SHEET_WEEKLY_DATA)

    dates = [
        row.get("Date")
        for row in rows
    ]

    values = [
        parse_number(row.get(series_name))
        for row in rows
    ]

    p80_values = rolling_percentile(values, rolling_window, 0.80)
    p20_values = rolling_percentile(values, rolling_window, 0.20)
    d4w_values = rolling_d4w(values)

    full_rows = []

    for current_date, value, p80, p20, d4w in zip(
        dates,
        values,
        p80_values,
        p20_values,
        d4w_values,
    ):
        full_rows.append(
            [
                current_date,
                value,
                p80,
                p20,
                None,
                d4w,
            ]
        )

    recent_rows = full_rows[-max_points:]
    latest_values = latest_marker_values([row[1] for row in recent_rows])

    for index, latest_value in enumerate(latest_values):
        recent_rows[index][4] = latest_value

    return recent_rows


def build_curve_rows(workbook, max_points: int = 260) -> list[list]:
    """
    Build 10Y-2Y curve slope chart block AL:AO.

    Output columns:
    Date, Curve 10Y-2Y, Zero Line, Latest Curve 10Y-2Y
    """
    rows = read_sheet_rows(workbook, SHEET_WEEKLY_DATA)

    full_rows = []

    for row in rows:
        dgs10 = parse_number(row.get("dgs10"))
        dgs2 = parse_number(row.get("dgs2"))

        if dgs10 is None or dgs2 is None:
            curve = None
        else:
            curve = dgs10 - dgs2

        full_rows.append(
            [
                row.get("Date"),
                curve,
                0,
                None,
            ]
        )

    recent_rows = full_rows[-max_points:]
    latest_values = latest_marker_values([row[1] for row in recent_rows])

    for index, latest_value in enumerate(latest_values):
        recent_rows[index][3] = latest_value

    return recent_rows


def build_stress_event_rows(workbook) -> list[list]:
    """
    Build stress-event snapshot block AQ:AW.
    """
    rows = read_sheet_rows(workbook, SHEET_STRESS_EVENT_AUDIT)

    output = []

    for row in rows:
        output.append(
            [
                row.get("Event"),
                row.get("Diagnostic Result"),
                row.get("Response Lag (Weeks)"),
                row.get("Peak Risk Appetite Regime"),
                row.get("Peak Rates & Inflation Regime"),
                row.get("Peak Drivers"),
                row.get("Pre-Existing Stress at Window Start"),
            ]
        )

    return output


def update_dashboard_data(output_path: str | Path) -> None:
    """
    Update _DASH_DATA in the output workbook.

    This function preserves the DASHBOARD sheet and its Excel-designed charts.
    """
    output_path = Path(output_path)
    workbook = load_workbook(output_path)

    if DASH_DATA_SHEET not in workbook.sheetnames:
        workbook.create_sheet(DASH_DATA_SHEET)

    ws = workbook[DASH_DATA_SHEET]

    clear_dash_data(ws)

    write_table(
        ws=ws,
        start_row=1,
        start_col=1,
        headers=["Metric", "Value"],
        rows=build_kpi_rows(workbook),
    )

    write_table(
        ws=ws,
        start_row=1,
        start_col=4,
        headers=[
            "Date",
            "Risk Appetite Score",
            "Risk Appetite Risk-On / Neutral Boundary",
            "Risk Appetite Neutral / Risk-Off Boundary",
            "Latest Risk Appetite Score",
            "Risk Appetite Regime Code",
        ],
        rows=build_risk_appetite_rows(workbook),
    )

    write_table(
        ws=ws,
        start_row=1,
        start_col=11,
        headers=[
            "Date",
            "Rates & Inflation Regime Code",
            "Latest Rates & Inflation Regime Code",
            "Rate Pressure Binary",
            "Inflation Concern Binary",
        ],
        rows=build_rates_inflation_timeline_rows(workbook),
    )

    write_table(
        ws=ws,
        start_row=1,
        start_col=17,
        headers=[
            "Date",
            "DGS10",
            "DGS10 52W P80 Threshold",
            "DGS10 52W P20 Context Band",
            "Latest DGS10",
            "DGS10 4W Change",
        ],
        rows=build_series_rows(workbook, "dgs10"),
    )

    write_table(
        ws=ws,
        start_row=1,
        start_col=24,
        headers=[
            "Date",
            "T10YIE",
            "T10YIE 52W P80 Threshold",
            "T10YIE 52W P20 Context Band",
            "Latest T10YIE",
            "T10YIE 4W Change",
        ],
        rows=build_series_rows(workbook, "t10yie"),
    )

    write_table(
        ws=ws,
        start_row=1,
        start_col=31,
        headers=[
            "Date",
            "DFII10",
            "DFII10 52W P80 Threshold",
            "DFII10 52W P20 Context Band",
            "Latest DFII10",
            "DFII10 4W Change",
        ],
        rows=build_series_rows(workbook, "dfii10"),
    )

    write_table(
        ws=ws,
        start_row=1,
        start_col=38,
        headers=[
            "Date",
            "10Y-2Y Curve Slope",
            "Zero Line",
            "Latest 10Y-2Y Curve Slope",
        ],
        rows=build_curve_rows(workbook),
    )

    write_table(
        ws=ws,
        start_row=1,
        start_col=43,
        headers=[
            "Event",
            "Diagnostic Result",
            "Response Lag (Weeks)",
            "Peak Risk Appetite Regime",
            "Peak Rates & Inflation Regime",
            "Peak Drivers",
            "Pre-Existing Stress at Window Start",
        ],
        rows=build_stress_event_rows(workbook),
    )

    # Keep visible during update. export.py hides it after this function runs.
    ws.sheet_state = "visible"

    remove_empty_dashboard_sheet(workbook)

    workbook.save(output_path)

    print("_DASH_DATA updated successfully.")