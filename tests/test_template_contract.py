from pathlib import Path

from openpyxl import load_workbook


def test_public_template_contract():
    template = (
        Path(__file__).resolve().parents[1]
        / "templates"
        / "dashboard_template.xlsx"
    )
    workbook = load_workbook(template, read_only=False, data_only=False)

    expected = {
        "DASHBOARD",
        "PDF_MEMO_1",
        "PDF_MEMO_2",
        "_DASH_DATA",
        "_PDF_DATA",
        "Stress Event Audit",
        "Regime Co-occurrence",
    }
    assert expected.issubset(set(workbook.sheetnames))
    assert len(workbook["DASHBOARD"]._charts) >= 1
    assert len(workbook["PDF_MEMO_2"]._charts) >= 1
