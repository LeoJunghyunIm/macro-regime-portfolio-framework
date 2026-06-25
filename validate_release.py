"""Validate the public repository before publishing it."""

from __future__ import annotations

import py_compile
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader

ROOT = Path(__file__).resolve().parent

REQUIRED_FILES = [
    "README.md",
    "LICENSE",
    ".gitignore",
    "requirements.txt",
    "config.yaml",
    "run_weekly.py",
    "risk_appetite_engine.py",
    "rates_inflation_engine.py",
    "portfolio_router.py",
    "quality_checks.py",
    "diagnostics.py",
    "dashboard.py",
    "pdf_memo.py",
    "export.py",
    "export_pdf.py",
    "templates/dashboard_template.xlsx",
    "sample_output/Macro_Regime_Framework_Memo_Sample.pdf",
]

TEXT_SUFFIXES = {".py", ".md", ".yaml", ".yml", ".txt"}


def _forbidden_terms() -> list[str]:
    # Construct retired development labels without displaying them in source.
    codes = (
        (87, 76, 49),
        (87, 76, 50),
        (87, 76, 51),
        (118, 49, 46, 48),
    )
    return ["".join(chr(value) for value in term) for term in codes]


def validate_required_files() -> list[str]:
    return [path for path in REQUIRED_FILES if not (ROOT / path).exists()]


def validate_source_text() -> list[str]:
    findings: list[str] = []
    forbidden = _forbidden_terms()
    privacy_markers = [
        "C:" + "\\" + "Users",
        "One" + "Drive",
        "im" + "jh1",
    ]

    for path in ROOT.rglob("*"):
        if not path.is_file() or path.suffix.lower() not in TEXT_SUFFIXES:
            continue
        if any(part in {".git", ".venv", "__pycache__"} for part in path.parts):
            continue

        text = path.read_text(encoding="utf-8", errors="ignore")
        for term in forbidden + privacy_markers:
            if term in text:
                findings.append(f"{path.relative_to(ROOT)} contains a retired or private marker")
                break

    return findings


def validate_python_syntax() -> list[str]:
    errors: list[str] = []
    for path in ROOT.glob("*.py"):
        try:
            py_compile.compile(str(path), doraise=True)
        except py_compile.PyCompileError as error:
            errors.append(str(error))
    return errors


def validate_template() -> list[str]:
    errors: list[str] = []
    path = ROOT / "templates" / "dashboard_template.xlsx"
    workbook = load_workbook(path, read_only=False, data_only=False)

    expected = {
        "DASHBOARD",
        "PDF_MEMO_1",
        "PDF_MEMO_2",
        "_DASH_DATA",
        "_PDF_DATA",
        "Stress Event Audit",
        "Regime Co-occurrence",
    }
    missing = expected.difference(workbook.sheetnames)
    if missing:
        errors.append(f"Template is missing sheets: {sorted(missing)}")

    if len(workbook["DASHBOARD"]._charts) < 1:
        errors.append("Dashboard charts are missing from the template")
    if len(workbook["PDF_MEMO_2"]._charts) < 1:
        errors.append("Memo charts are missing from the template")

    for sheet_name in workbook.sheetnames:
        if any(term in sheet_name for term in _forbidden_terms()):
            errors.append(f"Template contains a retired sheet label: {sheet_name}")

    return errors


def validate_sample_pdf() -> list[str]:
    path = ROOT / "sample_output" / "Macro_Regime_Framework_Memo_Sample.pdf"
    reader = PdfReader(path)
    if len(reader.pages) != 2:
        return [f"Sample memo must contain two pages; found {len(reader.pages)}"]
    return []


def main() -> None:
    checks = {
        "required files": validate_required_files(),
        "source text": validate_source_text(),
        "python syntax": validate_python_syntax(),
        "Excel template": validate_template(),
        "sample PDF": validate_sample_pdf(),
    }

    failed = False
    for name, errors in checks.items():
        if errors:
            failed = True
            print(f"FAIL - {name}")
            for error in errors:
                print(f"  - {error}")
        else:
            print(f"PASS - {name}")

    if failed:
        raise SystemExit(1)

    print("Release validation passed.")


if __name__ == "__main__":
    main()
