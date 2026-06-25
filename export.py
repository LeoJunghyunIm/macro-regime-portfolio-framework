"""
export.py

Export module for the Macro Regime-to-Portfolio Decision Framework.

Purpose:
- Export raw data, weekly aligned data, signal data, Risk Appetite,
  Rates & Inflation, Portfolio Router, data quality results, and diagnostics
  to Excel.
- Generate a plain-text Weekly Macro-to-Portfolio Memo.
- Archive weekly outputs by as-of date.
- Support template-based Excel dashboard workflow.

Important:
- Excel explains.
- Memo decides.
- DASHBOARD, PDF memo sheets, Stress Event Audit, and Regime Co-occurrence
  are human-designed in templates/dashboard_template.xlsx.
- Python updates values inside those designed sheets without replacing them.
- _DASH_DATA and _PDF_DATA are Python-updated and hidden after export.
"""

import shutil
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from dashboard import prepare_output_workbook_from_template, update_dashboard_data
from diagnostics import build_cooccurrence_matrix, build_stress_event_diagnostics
from pdf_memo import update_pdf_data


PUBLIC_SHEET_NAMES = {
    "risk_appetite": "Risk Appetite",
    "rates_inflation": "Rates & Inflation",
    "portfolio_router": "Portfolio Router",
    "data_quality": "Data Quality",
    "stress_event_audit": "Stress Event Audit",
    "regime_cooccurrence": "Regime Co-occurrence",
    "signal_data": "Signal Data",
    "weekly_data": "Weekly Data",
    "raw_data": "Raw Data",
}

# These are old generated sheets that may safely be removed from the copied
# template or from an old output. The two diagnostic template sheets are
# intentionally NOT included because their design must survive.
LEGACY_EXCEL_SHEETS = [
    "QA",
    "SIGNAL_DATA",
    "WEEKLY_DATA",
    "RAW_DATA",
]


INTERNAL_SHEETS_TO_HIDE = [
    "SUMMARY",
    "_DASH_DATA",
    "_PDF_DATA",
    "Signal Data",
    "Weekly Data",
    "Raw Data",
]

TEMPLATE_VALIDATION_SHEET_ALIASES = {
    "Stress Event Audit": ["Stress Event Audit"],
    "Regime Co-occurrence": ["Regime Co-occurrence"],
}


PUBLIC_COLUMN_RENAMES = {
    "AsOfDate": "As of Date",

    "Risk_Appetite_Valid": "Risk Appetite Valid",
    "Risk_Appetite_Regime": "Risk Appetite Regime",
    "Risk_Appetite_Score": "Risk Appetite Score",
    "Risk_Appetite_Trend": "Risk Appetite Trend",
    "Next_Risk_Appetite_Trigger": "Next Risk Appetite Trigger",

    "Rates_Inflation_Valid": "Rates & Inflation Valid",
    "Rates_Inflation_Regime": "Rates & Inflation Regime",
    "Next_Rates_Inflation_Trigger": "Next Rates & Inflation Trigger",

    "Portfolio_Router_Valid": "Portfolio Router Valid",
    "Portfolio_Adjustments": "Portfolio Adjustments",
    "Portfolio_Notes": "Portfolio Notes",

    "QA_Flag_Summary": "Data Quality Summary",

    "Pre_Risk_Appetite_Regime": "Pre-Event Risk Appetite Regime",
    "Pre_Rates_Inflation_Regime": "Pre-Event Rates & Inflation Regime",
    "Peak_Risk_Appetite_Regime": "Peak Risk Appetite Regime",
    "Peak_Risk_Appetite_Score": "Peak Risk Appetite Score",
    "Peak_Rates_Inflation_Regime": "Peak Rates & Inflation Regime",

    "Response_Lag_Weeks": "Response Lag (Weeks)",
    "Miss_Status": "Diagnostic Result",
    "Already_Stressed_At_Window_Start": "Pre-Existing Stress at Window Start",
    "First_Warning_Date": "First Warning Date",
    "Peak_Stress_Date": "Peak Stress Date",
    "Expected_Response": "Expected Response",
    "First_Warning_Rule": "First Warning Rule",
    "Framework_Valid": "Framework Valid",
    "Stress_Type": "Stress Type",
    "Window_Start": "Window Start",
    "Window_End": "Window End",
    "Peak_Drivers": "Peak Drivers",
    "Real_Yield_Tightening_Seen_In_Window": "Real Yield Tightening Seen in Window",
    "Real_Yield_Tightening_At_Peak": "Real Yield Tightening at Peak",
    "Diagnostic_Note": "Diagnostic Note",

    "check": "Check",
    "status": "Status",
    "details": "Details",
}

PUBLIC_TEXT_REPLACEMENTS = [
    ("Next_Risk_Appetite_Trigger", "Next Risk Appetite Trigger"),
    ("Next_Rates_Inflation_Trigger", "Next Rates & Inflation Trigger"),

    ("Risk_Appetite_Score", "Risk Appetite Score"),
    ("Risk_Appetite_Regime", "Risk Appetite Regime"),
    ("Risk_Appetite_Valid", "Risk Appetite Valid"),

    ("Rates_Inflation_Regime", "Rates & Inflation Regime"),
    ("Rates_Inflation_Valid", "Rates & Inflation Valid"),

    ("Portfolio_Adjustments", "Portfolio Adjustments"),
    ("Portfolio_Notes", "Portfolio Notes"),
    ("Portfolio_Router_Valid", "Portfolio Router Valid"),

    ("Pre_Risk_Appetite_Regime", "Pre-Event Risk Appetite Regime"),
    ("Pre_Rates_Inflation_Regime", "Pre-Event Rates & Inflation Regime"),
    ("Peak_Risk_Appetite_Regime", "Peak Risk Appetite Regime"),
    ("Peak_Risk_Appetite_Score", "Peak Risk Appetite Score"),
    ("Peak_Rates_Inflation_Regime", "Peak Rates & Inflation Regime"),

    (
        "Risk Appetite, Rates & Inflation and Portfolio Router",
        "Risk Appetite, Rates & Inflation, and Portfolio Router",
    ),
    ("Risk Appetite/Rates & Inflation", "Risk Appetite / Rates & Inflation"),


    ("Risk_Appetite_Trend", "Risk Appetite Trend"),
    ("AsOfDate", "As-of Date"),
    ("d4w", "4W Change"),
    ("QA", "Data Quality"),
]


def build_output_paths(config: dict) -> dict:
    """Build output file paths from config.yaml."""
    output_paths = {
        "excel_file": config["outputs"]["excel_file"],
        "memo_file": config["outputs"]["memo_file"],
    }

    output_directory = Path(config["outputs"]["directory"])
    output_directory.mkdir(parents=True, exist_ok=True)

    return output_paths


def replace_public_terms(value):
    """Replace internal labels with public-facing labels."""
    if not isinstance(value, str):
        return value

    output = value
    for old, new in PUBLIC_TEXT_REPLACEMENTS:
        output = output.replace(old, new)

    return output


def format_check_name(check_name) -> str:
    """Convert internal data-quality check names into readable labels."""
    if check_name is None:
        return "Unknown Check"

    text = str(check_name)

    mapping = {
        "missing_latest_value": "Missing Latest Value",
        "series_freshness": "Series Freshness",
        "rolling_window_sufficiency": "Rolling Window Sufficiency",
        "as_of_date_alignment": "As-of Date Alignment",
        "portfolio_output_available": "Portfolio Router Output Available",
        "latest_signal_validity": "Latest Signal Validity",
        "export_success": "Export Success",
        "pre_export_qa": "Pre-Export Data Quality",
    }

    if text in mapping:
        return mapping[text]

    text = replace_public_terms(text)
    text = text.replace("_", " ").strip().title()
    text = text.replace("As Of", "As-of")
    return text


def build_qa_flag_summary(qa_output: dict) -> str:
    """Build a concise Data Quality summary for the SUMMARY sheet."""
    qa_results = qa_output.get("qa_results", [])

    flagged_results = [
        result
        for result in qa_results
        if result.get("status") in ["WARN", "FAIL"]
    ]

    if not flagged_results:
        return "All pre-export data quality checks passed."

    summaries = []
    for result in flagged_results:
        check = format_check_name(result.get("check", "unknown_check"))
        status = result.get("status", "UNKNOWN")
        details = replace_public_terms(result.get("details", ""))
        summaries.append(f"{status} - {check}: {details}")

    return " | ".join(summaries)


def format_valid_status(value) -> str:
    """Format boolean validity flags for human-readable output."""
    return "VALID" if bool(value) else "INVALID"


def format_score_for_output(score, is_valid) -> str:
    """Format a regime score safely for Excel or memo output."""
    if not bool(is_valid):
        return "N/A"

    try:
        return str(int(score))
    except (TypeError, ValueError):
        return "N/A"


def format_on_off(value) -> str:
    """Format boolean-like values as ON/OFF."""
    return "ON" if bool(value) else "OFF"


def make_public_dataframe(dataframe: pd.DataFrame) -> pd.DataFrame:
    """Return a public-facing copy of a DataFrame."""
    public_df = dataframe.copy()
    public_df = public_df.rename(columns=PUBLIC_COLUMN_RENAMES)

    public_df.index.name = replace_public_terms(public_df.index.name)
    public_df.columns.name = replace_public_terms(public_df.columns.name)

    public_df = public_df.map(replace_public_terms)

    if "Check" in public_df.columns:
        public_df["Check"] = public_df["Check"].map(format_check_name)

    return public_df


def remove_legacy_excel_sheets(output_path: str | Path) -> None:
    """
    Remove obsolete generated sheets from a copied template workbook.

    The template validation sheets are not removed because their formatting
    must be preserved and renamed later.
    """
    path = Path(output_path)
    if not path.exists():
        return

    workbook = load_workbook(path)
    removed_any = False

    for sheet_name in LEGACY_EXCEL_SHEETS:
        if sheet_name in workbook.sheetnames:
            workbook.remove(workbook[sheet_name])
            removed_any = True

    if removed_any:
        workbook.save(path)


def _get_or_create_template_sheet(
    workbook,
    canonical_name: str,
    aliases: list[str],
):
    """
    Return a template worksheet without discarding its formatting.

    If an internal template name exists, rename that same worksheet to the
    public-facing name. Duplicate aliases are removed afterward.
    """
    if canonical_name in workbook.sheetnames:
        worksheet = workbook[canonical_name]
        created = False
    else:
        worksheet = None
        created = False

        for alias in aliases:
            if alias in workbook.sheetnames:
                worksheet = workbook[alias]
                worksheet.title = canonical_name
                break

        if worksheet is None:
            worksheet = workbook.create_sheet(canonical_name)
            created = True

    for alias in aliases:
        if (
            alias != canonical_name
            and alias in workbook.sheetnames
            and workbook[alias] is not worksheet
        ):
            workbook.remove(workbook[alias])

    return worksheet, created


def _clear_values_keep_styles(worksheet) -> None:
    """
    Clear cell values while preserving template formatting and page setup.
    """
    max_row = max(worksheet.max_row, 1)
    max_column = max(worksheet.max_column, 1)

    for row in worksheet.iter_rows(
        min_row=1,
        max_row=max_row,
        min_col=1,
        max_col=max_column,
    ):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def _excel_safe_value(value):
    """Convert pandas/NumPy values into openpyxl-safe values."""
    if value is None:
        return None

    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass

    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()

    if hasattr(value, "item"):
        try:
            return value.item()
        except (TypeError, ValueError):
            pass

    return value


def _write_dataframe_values(
    worksheet,
    dataframe: pd.DataFrame,
    start_row: int,
    start_column: int = 1,
    include_index: bool = False,
) -> tuple[int, int]:
    """
    Write DataFrame values directly into a styled worksheet.

    Unlike pandas if_sheet_exists='replace', this does not delete the sheet.
    """
    current_row = start_row
    current_column = start_column

    if include_index:
        index_cell = worksheet.cell(row=current_row, column=current_column)
        if not isinstance(index_cell, MergedCell):
            index_cell.value = _excel_safe_value(dataframe.index.name)
        header_start_column = current_column + 1
    else:
        header_start_column = current_column

    for column_offset, column_name in enumerate(dataframe.columns):
        cell = worksheet.cell(
            row=current_row,
            column=header_start_column + column_offset,
        )
        if not isinstance(cell, MergedCell):
            cell.value = _excel_safe_value(column_name)

    for row_offset, (index_value, row_values) in enumerate(
        dataframe.iterrows(),
        start=1,
    ):
        target_row = current_row + row_offset

        if include_index:
            index_cell = worksheet.cell(
                row=target_row,
                column=current_column,
            )
            if not isinstance(index_cell, MergedCell):
                index_cell.value = _excel_safe_value(index_value)
            value_start_column = current_column + 1
        else:
            value_start_column = current_column

        for column_offset, value in enumerate(row_values.tolist()):
            cell = worksheet.cell(
                row=target_row,
                column=value_start_column + column_offset,
            )
            if isinstance(cell, MergedCell):
                continue
            cell.value = _excel_safe_value(value)

    last_row = current_row + len(dataframe)
    last_column = (
        current_column
        + len(dataframe.columns)
        + (1 if include_index else 0)
        - 1
    )

    return last_row, last_column


def _apply_fallback_table_format(
    worksheet,
    table_header_rows: list[int],
) -> None:
    """Apply basic formatting only when a template sheet did not exist."""
    for row_number in table_header_rows:
        for cell in worksheet[row_number]:
            if cell.value is not None:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )

    for column_index in range(1, worksheet.max_column + 1):
        column_letter = get_column_letter(column_index)
        max_length = 0

        for row_number in range(1, worksheet.max_row + 1):
            value = worksheet.cell(
                row=row_number,
                column=column_index,
            ).value
            if value is not None:
                max_length = max(max_length, len(str(value)))

        worksheet.column_dimensions[column_letter].width = min(
            max(max_length + 2, 12),
            34,
        )


def update_validation_sheets_preserve_template(
    output_path: str | Path,
    stress_event_diagnostics: pd.DataFrame,
    percentage_matrix: pd.DataFrame,
) -> None:
    """
    Update public validation worksheets without replacing their design.

    Final public validation sheets:
    - Stress Event Audit
    - Regime Co-occurrence

    Regime Co-occurrence contains one percentage matrix only.
    """
    path = Path(output_path)
    if not path.exists():
        raise FileNotFoundError(f"Workbook not found: {path}")

    workbook = load_workbook(path)

    stress_name = PUBLIC_SHEET_NAMES["stress_event_audit"]
    cooccurrence_name = PUBLIC_SHEET_NAMES["regime_cooccurrence"]

    stress_ws, stress_created = _get_or_create_template_sheet(
        workbook,
        stress_name,
        TEMPLATE_VALIDATION_SHEET_ALIASES[stress_name],
    )
    cooccurrence_ws, cooccurrence_created = _get_or_create_template_sheet(
        workbook,
        cooccurrence_name,
        TEMPLATE_VALIDATION_SHEET_ALIASES[cooccurrence_name],
    )

    _clear_values_keep_styles(stress_ws)
    _clear_values_keep_styles(cooccurrence_ws)

    # Stress Event Audit: write the public table into A1:Y8 while retaining
    # the template's widths, fills, borders, row heights, and Y:AA merges.
    stress_last_row, stress_last_column = _write_dataframe_values(
        stress_ws,
        stress_event_diagnostics,
        start_row=1,
        start_column=1,
        include_index=False,
    )

    date_headers = {
        "Window Start",
        "Window End",
        "First Warning Date",
        "Peak Stress Date",
    }

    for column_number in range(1, stress_last_column + 1):
        header = stress_ws.cell(row=1, column=column_number).value
        if header in date_headers:
            for row_number in range(2, stress_last_row + 1):
                stress_ws.cell(
                    row=row_number,
                    column=column_number,
                ).number_format = "yyyy-mm-dd"

    # Regime Co-occurrence: one percentage matrix only, matching the
    # template's A1:F5 design.
    cooccurrence_ws["A1"] = "Regime Co-occurrence"

    for column_offset, column_name in enumerate(
        percentage_matrix.columns,
        start=2,
    ):
        cooccurrence_ws.cell(
            row=1,
            column=column_offset,
        ).value = _excel_safe_value(column_name)

    for row_offset, (index_value, row_values) in enumerate(
        percentage_matrix.iterrows(),
        start=2,
    ):
        cooccurrence_ws.cell(
            row=row_offset,
            column=1,
        ).value = _excel_safe_value(index_value)

        for column_offset, value in enumerate(
            row_values.tolist(),
            start=2,
        ):
            cell = cooccurrence_ws.cell(
                row=row_offset,
                column=column_offset,
            )
            cell.value = _excel_safe_value(value)

            if isinstance(cell.value, (int, float)) and 0 <= cell.value <= 1:
                cell.number_format = "0.0%"

    stress_ws.freeze_panes = stress_ws.freeze_panes or "B2"
    cooccurrence_ws.freeze_panes = cooccurrence_ws.freeze_panes or "B2"

    if stress_created:
        _apply_fallback_table_format(stress_ws, table_header_rows=[1])

    if cooccurrence_created:
        _apply_fallback_table_format(cooccurrence_ws, table_header_rows=[1])

    workbook.save(path)


def apply_public_sheet_order_and_visibility(output_path: str | Path) -> None:
    """Apply final public sheet order, visibility, and active sheet."""
    path = Path(output_path)
    if not path.exists():
        return

    workbook = load_workbook(path)

    desired_order = [
        "DASHBOARD",
        "PDF_MEMO_1",
        "PDF_MEMO_2",
        PUBLIC_SHEET_NAMES["portfolio_router"],
        PUBLIC_SHEET_NAMES["risk_appetite"],
        PUBLIC_SHEET_NAMES["rates_inflation"],
        PUBLIC_SHEET_NAMES["stress_event_audit"],
        PUBLIC_SHEET_NAMES["regime_cooccurrence"],
        PUBLIC_SHEET_NAMES["data_quality"],
        "SUMMARY",
        "_DASH_DATA",
        "_PDF_DATA",
        PUBLIC_SHEET_NAMES["signal_data"],
        PUBLIC_SHEET_NAMES["weekly_data"],
        PUBLIC_SHEET_NAMES["raw_data"],
    ]

    ordered_sheets = []
    used_names = set()

    for sheet_name in desired_order:
        if sheet_name in workbook.sheetnames:
            ordered_sheets.append(workbook[sheet_name])
            used_names.add(sheet_name)

    remaining_sheets = [
        worksheet
        for worksheet in workbook.worksheets
        if worksheet.title not in used_names
    ]

    workbook._sheets = ordered_sheets + remaining_sheets

    visible_public_sheets = {
        "DASHBOARD",
        "PDF_MEMO_1",
        "PDF_MEMO_2",
        PUBLIC_SHEET_NAMES["portfolio_router"],
        PUBLIC_SHEET_NAMES["risk_appetite"],
        PUBLIC_SHEET_NAMES["rates_inflation"],
        PUBLIC_SHEET_NAMES["stress_event_audit"],
        PUBLIC_SHEET_NAMES["regime_cooccurrence"],
        PUBLIC_SHEET_NAMES["data_quality"],
    }

    for worksheet in workbook.worksheets:
        if worksheet.title in INTERNAL_SHEETS_TO_HIDE:
            worksheet.sheet_state = "hidden"
        elif worksheet.title in visible_public_sheets:
            worksheet.sheet_state = "visible"

    if "DASHBOARD" in workbook.sheetnames:
        workbook.active = workbook.sheetnames.index("DASHBOARD")

    workbook.save(path)


def sanitize_workbook_text(output_path: str | Path) -> None:
    """Remove remaining internal labels from non-formula workbook text."""
    path = Path(output_path)
    if not path.exists():
        return

    workbook = load_workbook(path)

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                value = cell.value

                if not isinstance(value, str):
                    continue
                if value.startswith("="):
                    continue

                cleaned_value = replace_public_terms(value)
                if cleaned_value != value:
                    cell.value = cleaned_value

    workbook.save(path)


def format_excel_workbook(output_path: str | Path) -> None:
    """
    Apply minimal professional formatting to generated sheets.

    Human-designed dashboard, memo, and validation sheets are skipped.
    """
    workbook = load_workbook(output_path)

    manually_designed_sheets = {
        "DASHBOARD",
        "PDF_MEMO_1",
        "PDF_MEMO_2",
        PUBLIC_SHEET_NAMES["stress_event_audit"],
        PUBLIC_SHEET_NAMES["regime_cooccurrence"],
    }

    long_text_headers = {
        "Data Quality Summary",
        "Portfolio Adjustments",
        "Portfolio Notes",
        "Next Risk Appetite Trigger",
        "Next Rates & Inflation Trigger",
        "Real Yield Trigger",
        "Expected Response",
        "Diagnostic Note",
        "Details",
        "Text",
        "Value",
        "Peak Drivers",
    }

    date_headers = {
        "AsOfDate",
        "As of Date",
        "DATE",
        "Date",
        "date",
        "Start date",
        "End date",
    }

    for worksheet in workbook.worksheets:
        if worksheet.title in manually_designed_sheets:
            continue

        worksheet.freeze_panes = "A2"

        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        for column_index, column_cells in enumerate(
            worksheet.columns,
            start=1,
        ):
            column_letter = get_column_letter(column_index)
            header_value = worksheet.cell(
                row=1,
                column=column_index,
            ).value
            header_text = (
                str(header_value).strip()
                if header_value is not None
                else ""
            )

            max_length = 0
            for cell in column_cells:
                value = getattr(cell, "value", None)
                if value is None:
                    continue

                max_length = max(max_length, len(str(value)))

                if header_text in date_headers:
                    cell.number_format = "yyyy-mm-dd"

                if header_text in long_text_headers:
                    cell.alignment = Alignment(
                        wrap_text=True,
                        vertical="top",
                    )

            if header_text in long_text_headers:
                adjusted_width = 60
            elif header_text == "Field":
                adjusted_width = 34
            elif header_text in {"Value", "Text"}:
                adjusted_width = 80
            else:
                adjusted_width = min(max(max_length + 2, 12), 34)

            worksheet.column_dimensions[column_letter].width = adjusted_width

        if worksheet.title == "SUMMARY":
            worksheet.column_dimensions["A"].width = 34
            worksheet.column_dimensions["B"].width = 80

            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(
                        wrap_text=True,
                        vertical="top",
                    )

        if worksheet.title == PUBLIC_SHEET_NAMES["data_quality"]:
            worksheet.column_dimensions["A"].width = 32
            worksheet.column_dimensions["B"].width = 14
            worksheet.column_dimensions["C"].width = 90

            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(
                        wrap_text=True,
                        vertical="top",
                    )

        if worksheet.title == "_PDF_DATA":
            worksheet.column_dimensions["A"].width = 12
            worksheet.column_dimensions["B"].width = 32
            worksheet.column_dimensions["C"].width = 80
            worksheet.column_dimensions["E"].width = 16
            worksheet.column_dimensions["F"].width = 20
            worksheet.column_dimensions["G"].width = 58
            worksheet.column_dimensions["I"].width = 30
            worksheet.column_dimensions["J"].width = 20
            worksheet.column_dimensions["K"].width = 18
            worksheet.column_dimensions["L"].width = 24
            worksheet.column_dimensions["M"].width = 28
            worksheet.column_dimensions["N"].width = 54
            worksheet.column_dimensions["O"].width = 20
            worksheet.column_dimensions["Q"].width = 34
            worksheet.column_dimensions["R"].width = 90

            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(
                        wrap_text=True,
                        vertical="top",
                    )

    workbook.save(output_path)


def export_excel(
    raw_data,
    weekly_data,
    signal_data,
    risk_appetite_output,
    rates_inflation_output,
    portfolio_output,
    qa_output,
    output_path: str,
) -> None:
    """Export framework results to a presentation-ready Excel workbook."""
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    latest_date = portfolio_output.index.max()

    risk_appetite_latest = risk_appetite_output.loc[latest_date]
    rates_inflation_latest = rates_inflation_output.loc[latest_date]
    portfolio_latest = portfolio_output.loc[latest_date]

    confidence = qa_output.get("confidence", "Unknown")
    qa_flag_summary = build_qa_flag_summary(qa_output)

    risk_appetite_valid = bool(risk_appetite_latest.get("Risk_Appetite_Valid", False))
    rates_inflation_valid = bool(rates_inflation_latest.get("Rates_Inflation_Valid", False))
    portfolio_valid = bool(portfolio_latest.get("Portfolio_Router_Valid", False))

    real_yield_status = (
        format_on_off(portfolio_latest.get("Real_Yield_Tightening"))
        if portfolio_valid
        else "N/A"
    )

    summary = pd.DataFrame(
        {
            "Field": [
                "As of Date",
                "Risk Appetite Valid",
                "Risk Appetite Regime",
                "Risk Appetite Score",
                "Risk Appetite Trend",
                "Rates & Inflation Valid",
                "Rates & Inflation Regime",
                "Real Yield Tightening",
                "Portfolio Router Valid",
                "Confidence",
                "Data Quality Summary",
                "Equity Stance",
                "Duration Stance",
                "Cash Stance",
                "TIPS Stance",
                "Equity Band",
                "Duration Band",
                "Cash Band",
                "TIPS Band",
                "Portfolio Adjustments",
                "Portfolio Notes",
                "Next Risk Appetite Trigger",
                "Next Rates & Inflation Trigger",
                "Real Yield Trigger",
            ],
            "Value": [
                latest_date.date(),
                format_valid_status(risk_appetite_valid),
                risk_appetite_latest["Risk_Appetite_Regime"],
                format_score_for_output(
                    score=risk_appetite_latest.get("Risk_Appetite_Score"),
                    is_valid=risk_appetite_valid,
                ),
                risk_appetite_latest["Risk_Appetite_Trend"],
                format_valid_status(rates_inflation_valid),
                rates_inflation_latest["Rates_Inflation_Regime"],
                real_yield_status,
                format_valid_status(portfolio_valid),
                confidence,
                qa_flag_summary,
                portfolio_latest["Equity_Stance"],
                portfolio_latest["Duration_Stance"],
                portfolio_latest["Cash_Stance"],
                portfolio_latest["TIPS_Stance"],
                portfolio_latest["Equity_Band"],
                portfolio_latest["Duration_Band"],
                portfolio_latest["Cash_Band"],
                portfolio_latest["TIPS_Band"],
                replace_public_terms(portfolio_latest["Portfolio_Adjustments"]),
                replace_public_terms(portfolio_latest["Portfolio_Notes"]),
                replace_public_terms(risk_appetite_latest["Next_Risk_Appetite_Trigger"]),
                replace_public_terms(rates_inflation_latest["Next_Rates_Inflation_Trigger"]),
                replace_public_terms(rates_inflation_latest["Real_Yield_Trigger"]),
            ],
        }
    )

    qa_table = pd.DataFrame(qa_output.get("qa_results", []))
    qa_table = make_public_dataframe(qa_table)

    cooccurrence_output = build_cooccurrence_matrix(portfolio_output)
    cooccurrence_pct = make_public_dataframe(
        cooccurrence_output["pct_matrix"]
    )

    stress_event_diagnostics = build_stress_event_diagnostics(
        risk_appetite_output=risk_appetite_output,
        rates_inflation_output=rates_inflation_output,
        portfolio_output=portfolio_output,
    )
    stress_event_diagnostics = make_public_dataframe(
        stress_event_diagnostics
    )

    risk_appetite_output = make_public_dataframe(risk_appetite_output)
    rates_inflation_output = make_public_dataframe(rates_inflation_output)
    portfolio_router_output = make_public_dataframe(portfolio_output)

    # 1. Copy the designed template.
    template_used = prepare_output_workbook_from_template(path)

    if template_used:
        remove_legacy_excel_sheets(path)
        writer_kwargs = {
            "engine": "openpyxl",
            "mode": "a",
            "if_sheet_exists": "replace",
        }
    else:
        writer_kwargs = {"engine": "openpyxl"}

    # 2. Replace generated sheets only. The two validation template sheets
    # are excluded so their design survives.
    with pd.ExcelWriter(path, **writer_kwargs) as writer:
        summary.to_excel(
            writer,
            sheet_name="SUMMARY",
            index=False,
        )

        risk_appetite_output.to_excel(
            writer,
            sheet_name=PUBLIC_SHEET_NAMES["risk_appetite"],
        )
        rates_inflation_output.to_excel(
            writer,
            sheet_name=PUBLIC_SHEET_NAMES["rates_inflation"],
        )
        portfolio_router_output.to_excel(
            writer,
            sheet_name=PUBLIC_SHEET_NAMES["portfolio_router"],
        )

        qa_table.to_excel(
            writer,
            sheet_name=PUBLIC_SHEET_NAMES["data_quality"],
            index=False,
        )

        signal_data.to_excel(
            writer,
            sheet_name=PUBLIC_SHEET_NAMES["signal_data"],
        )
        weekly_data.to_excel(
            writer,
            sheet_name=PUBLIC_SHEET_NAMES["weekly_data"],
        )
        raw_data.to_excel(
            writer,
            sheet_name=PUBLIC_SHEET_NAMES["raw_data"],
        )

    # 3. Rename the template validation sheets and update values only.
    update_validation_sheets_preserve_template(
        output_path=path,
        stress_event_diagnostics=stress_event_diagnostics,
        percentage_matrix=cooccurrence_pct,
    )

    # 4. Format generated sheets while skipping the designed sheets.
    format_excel_workbook(path)

    # 5. Refresh dashboard and PDF data contracts.
    update_dashboard_data(str(path))
    update_pdf_data(str(path), keep_visible=False)

    # 6. Final cleanup and workbook presentation order.
    sanitize_workbook_text(path)
    apply_public_sheet_order_and_visibility(path)


def build_memo_text(
    risk_appetite_output,
    rates_inflation_output,
    portfolio_output,
    qa_output,
    config: dict,
) -> str:
    """Build the Weekly Macro-to-Portfolio Memo as plain text."""
    project_name = config["project"]["name"]
    latest_date = portfolio_output.index.max()

    risk_appetite_latest = risk_appetite_output.loc[latest_date]
    rates_inflation_latest = rates_inflation_output.loc[latest_date]
    portfolio_latest = portfolio_output.loc[latest_date]

    confidence = qa_output.get("confidence", "Unknown")

    risk_appetite_valid = bool(risk_appetite_latest.get("Risk_Appetite_Valid", False))
    rates_inflation_valid = bool(rates_inflation_latest.get("Rates_Inflation_Valid", False))
    portfolio_valid = bool(portfolio_latest.get("Portfolio_Router_Valid", False))

    risk_appetite_score_text = format_score_for_output(
        score=risk_appetite_latest.get("Risk_Appetite_Score"),
        is_valid=risk_appetite_valid,
    )

    real_yield_status = (
        format_on_off(portfolio_latest.get("Real_Yield_Tightening"))
        if portfolio_valid
        else "N/A"
    )

    risk_appetite_regime = risk_appetite_latest.get("Risk_Appetite_Regime", "N/A")
    rates_inflation_regime = rates_inflation_latest.get("Rates_Inflation_Regime", "N/A")
    risk_appetite_trend = risk_appetite_latest.get("Risk_Appetite_Trend", "N/A")

    equity_stance = portfolio_latest.get("Equity_Stance", "N/A")
    duration_stance = portfolio_latest.get("Duration_Stance", "N/A")
    cash_stance = portfolio_latest.get("Cash_Stance", "N/A")
    tips_stance = portfolio_latest.get("TIPS_Stance", "N/A")

    equity_band = portfolio_latest.get("Equity_Band", "N/A")
    duration_band = portfolio_latest.get("Duration_Band", "N/A")
    cash_band = portfolio_latest.get("Cash_Band", "N/A")
    tips_band = portfolio_latest.get("TIPS_Band", "N/A")

    data_quality_summary = build_qa_flag_summary(qa_output)

    if not risk_appetite_valid or not rates_inflation_valid or not portfolio_valid:
        return f"""\
{project_name}
Weekly Macro-to-Portfolio Memo

As of: {latest_date.date()}

1. Current Signal
Risk Appetite: {risk_appetite_regime}, Score: {risk_appetite_score_text}
Rates & Inflation Regime: {rates_inflation_regime}
Real Yield Tightening: {real_yield_status}
Confidence: {confidence}

2. Signal Validity
Risk Appetite Validity: {format_valid_status(risk_appetite_valid)}
Rates & Inflation Validity: {format_valid_status(rates_inflation_valid)}
Portfolio Router Validity: {format_valid_status(portfolio_valid)}

3. Interpretation
The latest framework output has insufficient data for at least one required layer.
Because the latest signal set is incomplete, the framework should not release a portfolio stance for this as-of date.

4. Portfolio Stance
Equity: N/A
Duration: N/A
Cash: N/A
TIPS: N/A

5. Next Trigger
Risk Appetite trigger: Insufficient data to generate Risk Appetite trigger.
Rates & Inflation trigger: Insufficient data to generate Rates & Inflation trigger.
Real Yield trigger: Insufficient data to generate Real Yield trigger.

6. Data Quality Summary
{data_quality_summary}

7. Key Risks and Limitations
- Latest signal validity failed or could not be confirmed.
- The framework should not be interpreted when required data is missing.
- Weekly percentile rules may lag fast market turning points.
- Signals are rule-based weekly regime indicators, not return forecasts.
- Stance bands are illustrative and not optimized target weights.
- The framework should be interpreted as a disciplined decision aid, not a trading system.
"""

    return f"""\
{project_name}
Weekly Macro-to-Portfolio Memo

As of: {latest_date.date()}

1. Current Signal
Risk Appetite: {risk_appetite_regime}, Score: {risk_appetite_score_text}
Risk Appetite Trend: {risk_appetite_trend}
Rates & Inflation Regime: {rates_inflation_regime}
Real Yield Tightening: {real_yield_status}
Confidence: {confidence}

2. Portfolio Stance
Equity: {equity_stance} ({equity_band})
Duration: {duration_stance} ({duration_band})
Cash: {cash_stance} ({cash_band})
TIPS: {tips_stance} ({tips_band})

3. Interpretation
Risk Appetite is classified as {risk_appetite_regime}, based on a score of {risk_appetite_score_text}.
The Rates & Inflation Regime is classified as {rates_inflation_regime}.
Real Yield Tightening is {real_yield_status}, which may cap aggressive equity or duration overweight positions when active.

The Portfolio Router translates the current macro regime into asset-allocation stance bands across equities, duration, cash, and TIPS.

4. What Changed
Portfolio Router notes: {replace_public_terms(portfolio_latest["Portfolio_Notes"])}

5. Next Trigger
Risk Appetite trigger: {replace_public_terms(risk_appetite_latest["Next_Risk_Appetite_Trigger"])}
Rates & Inflation trigger: {replace_public_terms(rates_inflation_latest["Next_Rates_Inflation_Trigger"])}
Real Yield trigger: {replace_public_terms(rates_inflation_latest["Real_Yield_Trigger"])}

6. Data Quality Summary
{data_quality_summary}

7. Key Risks and Limitations
- Weekly percentile rules may lag fast market turning points.
- Signals are rule-based weekly regime indicators, not return forecasts.
- Stance bands are illustrative and not optimized target weights.
- The framework should be interpreted as a disciplined decision aid, not a trading system.
"""


def export_weekly_memo(memo_text: str, output_path: str) -> None:
    """Export the weekly memo text to a .txt file."""
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(memo_text, encoding="utf-8")


def archive_output_files(output_paths: dict, as_of_date) -> dict:
    """Copy current outputs into a date-stamped local archive folder."""
    as_of_date_text = pd.Timestamp(as_of_date).date().isoformat()

    archive_dir = Path("outputs") / "archive" / as_of_date_text
    archive_dir.mkdir(parents=True, exist_ok=True)

    archive_paths = {}

    excel_source = Path(output_paths["excel_file"])
    memo_source = Path(output_paths["memo_file"])

    excel_archive = archive_dir / (
        f"Macro_Regime_Framework_Output_{as_of_date_text}.xlsx"
    )
    memo_archive = archive_dir / f"weekly_memo_{as_of_date_text}.txt"

    if excel_source.exists():
        shutil.copy2(excel_source, excel_archive)
        archive_paths["excel_archive"] = str(excel_archive)

    if memo_source.exists():
        shutil.copy2(memo_source, memo_archive)
        archive_paths["memo_archive"] = str(memo_archive)

    return archive_paths


def run_export(
    raw_data,
    weekly_data,
    signal_data,
    risk_appetite_output,
    rates_inflation_output,
    portfolio_output,
    qa_output,
    config: dict,
) -> dict:
    """Run the Excel, memo, and archive export steps."""
    output_paths = build_output_paths(config)

    export_excel(
        raw_data=raw_data,
        weekly_data=weekly_data,
        signal_data=signal_data,
        risk_appetite_output=risk_appetite_output,
        rates_inflation_output=rates_inflation_output,
        portfolio_output=portfolio_output,
        qa_output=qa_output,
        output_path=output_paths["excel_file"],
    )

    memo_text = build_memo_text(
        risk_appetite_output=risk_appetite_output,
        rates_inflation_output=rates_inflation_output,
        portfolio_output=portfolio_output,
        qa_output=qa_output,
        config=config,
    )

    export_weekly_memo(
        memo_text=memo_text,
        output_path=output_paths["memo_file"],
    )

    latest_date = portfolio_output.index.max()

    archive_paths = archive_output_files(
        output_paths=output_paths,
        as_of_date=latest_date,
    )

    output_paths.update(archive_paths)
    return output_paths

