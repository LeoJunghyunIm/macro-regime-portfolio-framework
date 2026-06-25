from __future__ import annotations

import re
import sys
from copy import copy
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook, workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PDF_DATA_SHEET = "_PDF_DATA"

SHEET_CANDIDATES = {
    "summary": ["SUMMARY"],
    "risk_appetite": ["Risk Appetite", "Risk Appetite_SIGNALS"],
    "rates_inflation": ["Rates & Inflation", "Rates & Inflation_SIGNALS"],
    "portfolio_router": ["Portfolio Router", "Portfolio Router_ROUTER"],
    "stress_event_audit": ["Stress Event Audit"],
}

TEXT_REPLACEMENTS = [
    ("Risk_Appetite_Score", "Risk Appetite Score"),
    ("Risk_Appetite_Regime", "Risk Appetite Regime"),
    ("Risk_Appetite_Valid", "Risk Appetite Valid"),
    ("Rates_Inflation_Regime", "Rates & Inflation Regime"),
    ("Rates_Inflation_Valid", "Rates & Inflation Valid"),
    ("Portfolio_Adjustments", "Portfolio Router Adjustments"),
    ("Portfolio_Notes", "Portfolio Router Notes"),
    ("Portfolio_Router_Valid", "Portfolio Router Valid"),
    ("Next_Risk_Appetite_Trigger", "Next Risk Appetite Trigger"),
    ("Next_Rates_Inflation_Trigger", "Next Rates & Inflation Trigger"),
    ("Risk_Appetite_Trend", "Risk Appetite Trend"),
    ("AsOfDate", "As of Date"),
    ("QA_Flag_Summary", "Data Quality Summary"),
    ("QA", "Data Quality"),
    ("Risk Appetite", "Risk Appetite"),
    ("Rates & Inflation", "Rates & Inflation"),
    ("Portfolio Router", "Portfolio Router"),
    ("d4w", "4W Change"),
]


def clean_public_text(value) -> str:
    """Return public-facing text for PDF memo output."""
    if value is None:
        return ""

    if isinstance(value, datetime):
        text = value.date().isoformat()
    elif isinstance(value, date):
        text = value.isoformat()
    else:
        text = str(value)

    for old, new in TEXT_REPLACEMENTS:
        text = text.replace(old, new)

    return text.strip()


def find_sheet(workbook, candidates: list[str]):
    for sheet_name in candidates:
        if sheet_name in workbook.sheetnames:
            return workbook[sheet_name]
    return None


def read_summary_fields(workbook) -> dict:
    worksheet = find_sheet(workbook, SHEET_CANDIDATES["summary"])
    if worksheet is None:
        return {}

    summary = {}

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue

        field = str(row[0]).strip()
        value = row[1] if len(row) > 1 else None
        summary[field] = value

    return summary


def get_summary_value(summary: dict, aliases: list[str], default=""):
    for alias in aliases:
        if alias in summary:
            return summary[alias]
    return default


def parse_regime_number(value):
    if value is None:
        return None

    if isinstance(value, (int, float)):
        return int(value)

    match = re.search(r"(\d+)", str(value))
    if match:
        return int(match.group(1))

    return None


def format_regime_label(regime_value) -> str:
    regime_num = parse_regime_number(regime_value)

    labels = {
        1: "Regime 1 — Rate + Inflation Pressure",
        2: "Regime 2 — Rate Pressure Only",
        3: "Regime 3 — Inflation Concern Only",
        4: "Regime 4 — No Rate / Inflation Pressure",
    }

    if regime_num in labels:
        return labels[regime_num]

    text = clean_public_text(regime_value)
    return text if text else "N/A"


def format_rates_inflation_display(regime_value) -> str:
    """
    Short display label for PDF memo.

    Public-facing examples:
    - Rate + Inflation Pressure
    - Rate Pressure Only
    - Inflation Concern Only
    - No Rate / Inflation Pressure
    """
    regime_num = parse_regime_number(regime_value)

    labels = {
        1: "Rate + Inflation Pressure",
        2: "Rate Pressure Only",
        3: "Inflation Concern Only",
        4: "No Rate / Inflation Pressure",
    }

    if regime_num in labels:
        return labels[regime_num]

    text = clean_public_text(regime_value)
    return text if text else "N/A"


def infer_pressure_flags(regime_value) -> tuple[str, str]:
    regime_num = parse_regime_number(regime_value)

    if regime_num == 1:
        return "ON", "ON"
    if regime_num == 2:
        return "ON", "OFF"
    if regime_num == 3:
        return "OFF", "ON"
    if regime_num == 4:
        return "OFF", "OFF"

    return "N/A", "N/A"


def format_on_off(value) -> str:
    if value is None:
        return "N/A"

    if isinstance(value, str):
        text = value.strip().upper()

        if text in {"ON", "TRUE", "YES", "1"}:
            return "ON"
        if text in {"OFF", "FALSE", "NO", "0"}:
            return "OFF"
        if text in {"N/A", "NA", ""}:
            return "N/A"

        return clean_public_text(value)

    return "ON" if bool(value) else "OFF"


def format_score(value) -> str:
    if value is None:
        return "N/A"

    if isinstance(value, (int, float)):
        return str(int(value)) if float(value).is_integer() else str(value)

    text = str(value).strip()

    try:
        number = float(text)
        return str(int(number)) if number.is_integer() else text
    except ValueError:
        return clean_public_text(text)


def format_date(value) -> str:
    if value is None:
        return "N/A"

    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, date):
        return value.isoformat()

    return clean_public_text(value)


def date_sort_key(value):
    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, date):
        return value.isoformat()

    return str(value)


def row_date_text(row: dict | None) -> str:
    if not row:
        return ""

    return format_date(row.get("__date__"))


def read_rows_by_date(workbook, sheet_key: str) -> list[dict]:
    worksheet = find_sheet(workbook, SHEET_CANDIDATES[sheet_key])
    if worksheet is None:
        return []

    headers = []

    for cell in worksheet[1]:
        headers.append(str(cell.value).strip() if cell.value is not None else "")

    rows = []

    for row_number in range(2, worksheet.max_row + 1):
        raw_date = worksheet.cell(row=row_number, column=1).value

        if raw_date is None:
            continue

        row_data = {
            "__date__": raw_date,
            "__row_number__": row_number,
        }

        for col_number, header in enumerate(headers, start=1):
            if not header:
                continue

            row_data[header] = worksheet.cell(row=row_number, column=col_number).value

        rows.append(row_data)

    rows.sort(key=lambda item: date_sort_key(item.get("__date__")))
    return rows


def row_value(row: dict | None, aliases: list[str], default=""):
    if not row:
        return default

    for alias in aliases:
        if alias in row:
            return row[alias]

    return default


def build_snapshot(workbook) -> tuple[dict, dict]:
    summary = read_summary_fields(workbook)

    risk_rows = read_rows_by_date(workbook, "risk_appetite")
    rates_rows = read_rows_by_date(workbook, "rates_inflation")
    router_rows = read_rows_by_date(workbook, "portfolio_router")

    current_risk = risk_rows[-1] if risk_rows else None
    previous_risk = risk_rows[-2] if len(risk_rows) >= 2 else None

    current_rates = rates_rows[-1] if rates_rows else None
    previous_rates = rates_rows[-2] if len(rates_rows) >= 2 else None

    current_router = router_rows[-1] if router_rows else None
    previous_router = router_rows[-2] if len(router_rows) >= 2 else None

    def is_missing(value) -> bool:
        if value is None:
            return True
        if isinstance(value, str):
            return value.strip() == "" or value.strip().upper() in {
                "N/A",
                "NA",
                "NONE",
                "NULL",
            }
        return False

    def clean_reason(value, default: str = "None") -> str:
        if is_missing(value):
            return default
        cleaned = clean_public_text(value)
        return cleaned if cleaned else default

    as_of_date = get_summary_value(
        summary,
        ["As of Date", "AsOfDate"],
        row_date_text(current_router) or row_date_text(current_risk) or "N/A",
    )

    risk_regime = get_summary_value(
        summary,
        ["Risk Appetite Regime", "Risk_Appetite_Regime"],
        row_value(current_risk, ["Risk Appetite Regime", "Risk_Appetite_Regime"], "N/A"),
    )

    risk_score = get_summary_value(
        summary,
        ["Risk Appetite Score", "Risk_Appetite_Score"],
        row_value(current_risk, ["Risk Appetite Score", "Risk_Appetite_Score"], "N/A"),
    )

    risk_trend = get_summary_value(
        summary,
        ["Risk Appetite Trend", "Risk_Appetite_Trend"],
        row_value(current_risk, ["Risk Appetite Trend", "Risk_Appetite_Trend"], "N/A"),
    )

    rates_regime = get_summary_value(
        summary,
        ["Rates & Inflation Regime", "Rates_Inflation_Regime"],
        row_value(current_rates, ["Rates & Inflation Regime", "Rates_Inflation_Regime"], "N/A"),
    )

    # Prefer SUMMARY values if present, but fall back to the Rates & Inflation sheet.
    rate_pressure = get_summary_value(
        summary,
        ["Rate Pressure"],
        row_value(current_rates, ["Rate_Pressure", "Rate Pressure"], ""),
    )

    inflation_concern = get_summary_value(
        summary,
        ["Inflation Concern"],
        row_value(current_rates, ["Inflation_Concern", "Inflation Concern"], ""),
    )

    if is_missing(rate_pressure) or is_missing(inflation_concern):
        inferred_rate, inferred_inflation = infer_pressure_flags(rates_regime)

        if is_missing(rate_pressure):
            rate_pressure = inferred_rate

        if is_missing(inflation_concern):
            inflation_concern = inferred_inflation

    real_yield_tightening = get_summary_value(
        summary,
        ["Real Yield Tightening", "Real_Yield_Tightening"],
        row_value(
            current_rates,
            ["Real_Yield_Tightening", "Real Yield Tightening"],
            "N/A",
        ),
    )

    confidence = get_summary_value(summary, ["Confidence"], "Unknown")

    data_quality = get_summary_value(
        summary,
        ["Data Quality Summary", "QA_Flag_Summary"],
        "N/A",
    )

    data_quality_text = clean_public_text(data_quality)

    if (
        "passed" in data_quality_text.lower()
        and "fail" not in data_quality_text.lower()
        and "warn" not in data_quality_text.lower()
    ):
        data_quality_display = "Pass"
    else:
        data_quality_display = data_quality_text if data_quality_text else "N/A"

    snapshot = {
        "as_of_date": format_date(as_of_date),

        # Risk Appetite headline fields
        "risk_regime": clean_public_text(risk_regime) or "N/A",
        "risk_score": format_score(risk_score),
        "risk_trend": clean_public_text(risk_trend) or "N/A",

        # Risk Appetite driver fields
        "credit_flag": row_value(current_risk, ["Credit_Flag", "Credit Flag"], ""),
        "hy_reason": clean_reason(
            row_value(current_risk, ["HY_Reason", "HY Reason"], "None")
        ),
        "nfci_flag": row_value(current_risk, ["NFCI_Flag", "NFCI Flag"], ""),
        "vix_flag": row_value(current_risk, ["VIX_Flag", "VIX Flag"], ""),
        "claims_flag": row_value(current_risk, ["Claims_Flag", "Claims Flag"], ""),

        # Rates & Inflation headline fields
        "rates_regime": clean_public_text(rates_regime) or "N/A",
        "rates_regime_label": format_regime_label(rates_regime),
        "rates_inflation_display": format_rates_inflation_display(rates_regime),
        "rate_pressure": format_on_off(rate_pressure),
        "inflation_concern": format_on_off(inflation_concern),
        "real_yield_tightening": format_on_off(real_yield_tightening),

        # Rates & Inflation driver fields
        "rate_reason": clean_reason(
            row_value(current_rates, ["Rate_Reason", "Rate Reason"], "None")
        ),
        "inflation_reason": clean_reason(
            row_value(current_rates, ["Inflation_Reason", "Inflation Reason"], "None")
        ),
        "real_yield_reason": clean_reason(
            row_value(current_rates, ["Real_Yield_Reason", "Real Yield Reason"], "None")
        ),

        # Quality / confidence
        "confidence": clean_public_text(confidence) or "Unknown",
        "data_quality": data_quality_display,

        # Portfolio Router fields
        "equity_stance": clean_public_text(
            get_summary_value(
                summary,
                ["Equity Stance", "Equity_Stance"],
                row_value(current_router, ["Equity Stance", "Equity_Stance"], "N/A"),
            )
        ) or "N/A",
        "duration_stance": clean_public_text(
            get_summary_value(
                summary,
                ["Duration Stance", "Duration_Stance"],
                row_value(current_router, ["Duration Stance", "Duration_Stance"], "N/A"),
            )
        ) or "N/A",
        "cash_stance": clean_public_text(
            get_summary_value(
                summary,
                ["Cash Stance", "Cash_Stance"],
                row_value(current_router, ["Cash Stance", "Cash_Stance"], "N/A"),
            )
        ) or "N/A",
        "tips_stance": clean_public_text(
            get_summary_value(
                summary,
                ["TIPS Stance", "TIPS_Stance"],
                row_value(current_router, ["TIPS Stance", "TIPS_Stance"], "N/A"),
            )
        ) or "N/A",
        "equity_band": clean_public_text(
            get_summary_value(
                summary,
                ["Equity Band", "Equity_Band"],
                row_value(current_router, ["Equity Band", "Equity_Band"], "N/A"),
            )
        ) or "N/A",
        "duration_band": clean_public_text(
            get_summary_value(
                summary,
                ["Duration Band", "Duration_Band"],
                row_value(current_router, ["Duration Band", "Duration_Band"], "N/A"),
            )
        ) or "N/A",
        "cash_band": clean_public_text(
            get_summary_value(
                summary,
                ["Cash Band", "Cash_Band"],
                row_value(current_router, ["Cash Band", "Cash_Band"], "N/A"),
            )
        ) or "N/A",
        "tips_band": clean_public_text(
            get_summary_value(
                summary,
                ["TIPS Band", "TIPS_Band"],
                row_value(current_router, ["TIPS Band", "TIPS_Band"], "N/A"),
            )
        ) or "N/A",
    }

    previous = {
        "as_of_date": row_date_text(previous_router) or row_date_text(previous_risk),

        # Previous Risk Appetite headline fields
        "risk_regime": clean_public_text(
            row_value(previous_risk, ["Risk Appetite Regime", "Risk_Appetite_Regime"], "")
        ),
        "risk_score": format_score(
            row_value(previous_risk, ["Risk Appetite Score", "Risk_Appetite_Score"], "")
        ),

        # Previous Risk Appetite driver fields
        "credit_flag": row_value(previous_risk, ["Credit_Flag", "Credit Flag"], ""),
        "hy_reason": clean_reason(
            row_value(previous_risk, ["HY_Reason", "HY Reason"], "None")
        ),
        "nfci_flag": row_value(previous_risk, ["NFCI_Flag", "NFCI Flag"], ""),
        "vix_flag": row_value(previous_risk, ["VIX_Flag", "VIX Flag"], ""),
        "claims_flag": row_value(previous_risk, ["Claims_Flag", "Claims Flag"], ""),

        # Previous Rates & Inflation headline fields
        "rates_regime": clean_public_text(
            row_value(previous_rates, ["Rates & Inflation Regime", "Rates_Inflation_Regime"], "")
        ),
        "rate_pressure": format_on_off(
            row_value(previous_rates, ["Rate_Pressure", "Rate Pressure"], "")
        ),
        "inflation_concern": format_on_off(
            row_value(previous_rates, ["Inflation_Concern", "Inflation Concern"], "")
        ),
        "real_yield_tightening": format_on_off(
            row_value(
                previous_rates,
                ["Real_Yield_Tightening", "Real Yield Tightening"],
                "",
            )
        ),

        # Previous Rates & Inflation driver fields
        "rate_reason": clean_reason(
            row_value(previous_rates, ["Rate_Reason", "Rate Reason"], "None")
        ),
        "inflation_reason": clean_reason(
            row_value(previous_rates, ["Inflation_Reason", "Inflation Reason"], "None")
        ),
        "real_yield_reason": clean_reason(
            row_value(previous_rates, ["Real_Yield_Reason", "Real Yield Reason"], "None")
        ),

        # Previous Portfolio Router fields
        "equity_stance": clean_public_text(
            row_value(previous_router, ["Equity Stance", "Equity_Stance"], "")
        ),
        "duration_stance": clean_public_text(
            row_value(previous_router, ["Duration Stance", "Duration_Stance"], "")
        ),
        "cash_stance": clean_public_text(
            row_value(previous_router, ["Cash Stance", "Cash_Stance"], "")
        ),
        "tips_stance": clean_public_text(
            row_value(previous_router, ["TIPS Stance", "TIPS_Stance"], "")
        ),
    }

    if is_missing(previous.get("rate_pressure")) or is_missing(previous.get("inflation_concern")):
        previous_rate, previous_inflation = infer_pressure_flags(previous.get("rates_regime"))

        if is_missing(previous.get("rate_pressure")):
            previous["rate_pressure"] = previous_rate

        if is_missing(previous.get("inflation_concern")):
            previous["inflation_concern"] = previous_inflation

    return snapshot, previous


def build_decision_summary(snapshot: dict) -> str:
    if snapshot["risk_regime"] == "Risk-On":
        risk_part = "Risk appetite remains constructive"
    elif snapshot["risk_regime"] == "Risk-Off":
        risk_part = "Risk appetite is defensive"
    else:
        risk_part = "Risk appetite is neutral"

    if snapshot["real_yield_tightening"] == "ON":
        constraint = "real-yield tightening limits aggressive risk-taking"
    elif snapshot["real_yield_tightening"] == "OFF":
        constraint = "the real-yield constraint is not active"
    else:
        constraint = "real-yield conditions require review"

    return f"{risk_part}, while {constraint}."


def build_executive_summary(snapshot: dict) -> str:
    portfolio_line = (
        f"Portfolio: Equity {snapshot['equity_stance']}, "
        f"Duration {snapshot['duration_stance']}, "
        f"Cash {snapshot['cash_stance']}, "
        f"TIPS {snapshot['tips_stance']}."
    )

    if snapshot["risk_regime"] == "Risk-On":
        risk_line = "Risk Appetite: Risk-On — no broad risk-stress signals are active."
    elif snapshot["risk_regime"] == "Risk-Off":
        risk_line = "Risk Appetite: Risk-Off — broad market-stress signals are active."
    else:
        risk_line = (
            f"Risk Appetite: {snapshot['risk_regime']} — risk conditions are neither fully constructive nor broadly defensive."
        )

    if snapshot["inflation_concern"] == "ON":
        inflation_text = "inflation concern is elevated"
    elif snapshot["inflation_concern"] == "OFF":
        inflation_text = "inflation concern is not elevated"
    else:
        inflation_text = "inflation concern requires review"

    rates_line = (
        f"Rates & Inflation: {snapshot['rates_inflation_display']} — {inflation_text}."
    )

    if snapshot["real_yield_tightening"] == "ON":
        real_yield_line = (
            "Real Yield Tightening: ON — aggressive equity and duration overweights remain constrained."
        )
    elif snapshot["real_yield_tightening"] == "OFF":
        real_yield_line = (
            "Real Yield Tightening: OFF — the real-yield constraint is not currently active."
        )
    else:
        real_yield_line = (
            "Real Yield Tightening: N/A — portfolio constraints require data review."
        )

    data_quality_line = f"Data Quality: {snapshot['data_quality']}."

    return "\n".join(
        [
            portfolio_line,
            risk_line,
            rates_line,
            real_yield_line,
            data_quality_line,
        ]
    )


def build_executive_interpretation(snapshot: dict) -> str:
    risk_regime = snapshot["risk_regime"]

    if risk_regime == "Risk-On":
        risk_sentence = "Risk appetite remains constructive, with no broad-based stress signal currently active."
    elif risk_regime == "Risk-Off":
        risk_sentence = "Risk appetite is defensive, indicating broad-based market stress across the signal set."
    else:
        risk_sentence = "Risk appetite is neutral, indicating neither a clean risk-on backdrop nor a broad stress regime."

    rates_sentence = (
        f"The rates and inflation backdrop is classified as {snapshot['rates_regime_label']}, "
        f"with Rate Pressure {snapshot['rate_pressure']} and Inflation Concern {snapshot['inflation_concern']}."
    )

    if snapshot["real_yield_tightening"] == "ON":
        real_yield_sentence = (
            "Real Yield Tightening is active, limiting the case for aggressive equity "
            "or duration overweight positions."
        )
    elif snapshot["real_yield_tightening"] == "OFF":
        real_yield_sentence = (
            "Real Yield Tightening is not active, so the portfolio constraint from real yields "
            "is less restrictive."
        )
    else:
        real_yield_sentence = (
            "Real Yield Tightening status is unavailable and should be reviewed before making "
            "allocation decisions."
        )

    portfolio_sentence = (
        "The Portfolio Router maps this backdrop to "
        f"Equity {snapshot['equity_stance']}, Duration {snapshot['duration_stance']}, "
        f"Cash {snapshot['cash_stance']}, and TIPS {snapshot['tips_stance']}."
    )

    return "\n\n".join(
        [
            risk_sentence,
            rates_sentence,
            real_yield_sentence,
            portfolio_sentence,
        ]
    )


def describe_change(label: str, previous_value: str, current_value: str) -> str:
    previous_text = clean_public_text(previous_value)
    current_text = clean_public_text(current_value)

    if not previous_text or previous_text == "N/A":
        return f"{label} prior-week value is unavailable."

    if previous_text == current_text:
        return f"{label} remained {current_text}."

    return f"{label} shifted from {previous_text} to {current_text}."


def build_change_summary(snapshot: dict, previous: dict) -> str:
    lines = [
        describe_change("Risk Appetite", previous.get("risk_regime"), snapshot.get("risk_regime")),
        describe_change("Risk Appetite Score", previous.get("risk_score"), snapshot.get("risk_score")),
        describe_change("Rates & Inflation", previous.get("rates_regime"), snapshot.get("rates_regime")),
        describe_change("Real Yield Tightening", previous.get("real_yield_tightening"), snapshot.get("real_yield_tightening")),
        describe_change("Equity stance", previous.get("equity_stance"), snapshot.get("equity_stance")),
        describe_change("Duration stance", previous.get("duration_stance"), snapshot.get("duration_stance")),
        describe_change("Cash stance", previous.get("cash_stance"), snapshot.get("cash_stance")),
        describe_change("TIPS stance", previous.get("tips_stance"), snapshot.get("tips_stance")),
    ]

    return "\n\n".join(lines)


def build_triggers(snapshot: dict) -> dict:
    try:
        score = int(float(snapshot.get("risk_score", "N/A")))
    except (TypeError, ValueError):
        score = None

    if score == 0:
        risk_trigger = (
            "A Risk Appetite Score of 1 would shift the framework from Risk-On toward Neutral."
        )
    elif score in {1, 2}:
        risk_trigger = (
            "A Risk Appetite Score of 3+ would shift the framework to Risk-Off; "
            "a score of 0 would restore Risk-On."
        )
    elif score is not None and score >= 3:
        risk_trigger = (
            "A Risk Appetite Score of 2 or lower would shift the framework back toward Neutral."
        )
    else:
        risk_trigger = (
            "Score availability should be reviewed before interpreting "
            "the next regime trigger."
        )

    regime_num = parse_regime_number(snapshot.get("rates_regime"))

    if regime_num == 1:
        rates_trigger = (
            "Easing either Rate Pressure or Inflation Concern would move "
            "the regime away from the most restrictive rate-and-inflation backdrop."
        )
    elif regime_num == 2:
        rates_trigger = (
            "If Inflation Concern turns ON while Rate Pressure remains ON, "
            "the regime shifts from Regime 2 to Regime 1."
        )
    elif regime_num == 3:
        rates_trigger = (
            "If Rate Pressure turns ON while Inflation Concern remains ON, "
            "the regime shifts from Regime 3 to Regime 1."
        )
    elif regime_num == 4:
        rates_trigger = (
            "Renewed Rate Pressure or Inflation Concern would move the "
            "framework out of Regime 4."
        )
    else:
        rates_trigger = (
            "Regime availability should be reviewed before interpreting "
            "the next trigger."
        )

    if snapshot.get("real_yield_tightening") == "ON":
        real_yield_trigger = (
            "While Real Yield Tightening remains ON, aggressive equity and duration "
            "overweights remain constrained."
        )
        portfolio_trigger = (
            "The main release valve would be Real Yield Tightening turning "
            "OFF or risk appetite deteriorating enough to justify a more defensive stance."
        )
    elif snapshot.get("real_yield_tightening") == "OFF":
        real_yield_trigger = (
            "If Real Yield Tightening turns ON, the Portfolio Router should cap "
            "aggressive equity or duration overweight positions."
        )
        portfolio_trigger = (
            "Without the real-yield constraint, portfolio stance depends "
            "more directly on the Risk Appetite and Rates & Inflation regimes."
        )
    else:
        real_yield_trigger = (
            "Status should be reviewed before interpreting portfolio constraints."
        )
        portfolio_trigger = (
            "Allocation signals should be treated with caution until data quality is confirmed."
        )

    trigger_summary = "\n\n".join(
        [
            risk_trigger,
            rates_trigger,
            real_yield_trigger,
            portfolio_trigger,
        ]
    )

    return {
        "risk_appetite_trigger": risk_trigger,
        "rates_inflation_trigger": rates_trigger,
        "real_yield_trigger": real_yield_trigger,
        "portfolio_constraint_trigger": portfolio_trigger,
        "trigger_summary": trigger_summary,
    }


def build_key_risks_and_limitations() -> str:
    return "\n\n".join(
        [
            "Weekly rules may lag fast market turning points.",
            "Signals are regime indicators, not return forecasts.",
            "Stance bands are illustrative and not optimized target weights.",
            "The framework is a disciplined decision aid, not a trading system.",
            "Real-time interpretation should consider market context beyond the rule-based signal set.",
        ]
    )


def build_rule_rationale() -> str:
    return "\n\n".join(
        [
            "Risk Appetite monitors credit, financial conditions, volatility, and labor stress.",
            "Rates & Inflation tracks rate pressure and inflation concern.",
            "Real Yield Tightening acts as a portfolio constraint, not a standalone bullish or bearish signal.",
            "The Portfolio Router maps regime combinations into asset-class stance bands.",
        ]
    )


def build_limitations() -> str:
    return "\n\n".join(
        [
            "Weekly signals may lag rapid market turning points.",
            "Percentile thresholds are designed for interpretability, not optimization.",
            "Stress-event diagnostics are framework audits, not return backtests.",
            "Portfolio stance bands are illustrative decision ranges.",
        ]
    )


def build_evidence_summaries(snapshot: dict) -> dict:
    risk_summary = (
        f"Risk Appetite Score is {snapshot['risk_score']}, consistent with a "
        f"{snapshot['risk_regime']} regime."
    )

    rates_summary = (
        f"Rates & Inflation is {snapshot['rates_regime_label']}; "
        f"Rate Pressure is {snapshot['rate_pressure']} and Inflation Concern is {snapshot['inflation_concern']}."
    )

    real_yield_summary = (
        f"Real Yield Tightening is {snapshot['real_yield_tightening']}, "
        "which determines whether aggressive equity and duration overweights are constrained."
    )

    portfolio_summary = (
        f"The Portfolio Router currently maps the macro backdrop to Equity {snapshot['equity_stance']}, "
        f"Duration {snapshot['duration_stance']}, Cash {snapshot['cash_stance']}, and TIPS {snapshot['tips_stance']}."
    )

    chart_summary = (
        "The chart evidence links the current risk backdrop, rates-and-inflation regime, "
        "real-yield constraint, and portfolio stance into a repeatable allocation framework."
    )

    return {
        "risk_appetite_evidence": risk_summary,
        "rates_inflation_evidence": rates_summary,
        "real_yield_evidence": real_yield_summary,
        "portfolio_evidence": portfolio_summary,
        "chart_evidence": chart_summary,
    }


def build_risk_appetite_driver(snapshot: dict, previous: dict) -> str:
    """
    Create a public-facing driver phrase for Risk Appetite changes.

    Uses actual Risk Appetite driver fields:
        - Credit_Flag + HY_Reason
        - NFCI_Flag
        - VIX_Flag
        - Claims_Flag

    Output is designed to fit inside Page 1 wording:
        "Moved from Risk-On to Neutral as {driver}."
    """

    def to_bool(value) -> bool:
        if value is None:
            return False

        if isinstance(value, bool):
            return value

        if isinstance(value, (int, float)):
            return value != 0

        if isinstance(value, str):
            cleaned = value.strip().lower()
            if cleaned in {"1", "true", "yes", "y", "on"}:
                return True
            if cleaned in {"0", "false", "no", "n", "off", "", "none", "n/a", "na"}:
                return False

        return False

    def to_score(value):
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return None

    def normalize_reason(value) -> str:
        if value is None:
            return "none"

        cleaned = str(value).strip().lower()

        if cleaned in {"", "none", "n/a", "na", "null"}:
            return "none"

        if cleaned in {"both", "level + momentum", "level and momentum", "level/momentum"}:
            return "both"

        if "both" in cleaned:
            return "both"

        if "momentum" in cleaned or "trend" in cleaned or "4w" in cleaned or "d4w" in cleaned:
            return "momentum"

        if "level" in cleaned or "p80" in cleaned:
            return "level"

        return cleaned

    def credit_components(source: dict) -> tuple[bool, bool]:
        """
        Returns:
            credit_level_active, credit_momentum_active
        """
        credit_flag = to_bool(source.get("credit_flag"))
        reason = normalize_reason(source.get("hy_reason"))

        if not credit_flag:
            return False, False

        if reason == "both":
            return True, True

        if reason == "level":
            return True, False

        if reason == "momentum":
            return False, True

        # Conservative fallback:
        # If Credit_Flag is ON but HY_Reason is unclear, treat it as generic credit stress.
        return True, False

    def join_two(phrases: list[str]) -> str:
        if not phrases:
            return ""
        if len(phrases) == 1:
            return phrases[0]
        if len(phrases) == 2:
            return f"{phrases[0]} and {phrases[1]}"
        return f"{', '.join(phrases[:-1])}, and {phrases[-1]}"

    def format_channels(channels: list[str]) -> str:
        unique = []
        for channel in channels:
            if channel not in unique:
                unique.append(channel)

        if len(unique) == 1:
            return unique[0]
        if len(unique) == 2:
            return f"{unique[0]} and {unique[1]}"
        return f"{', '.join(unique[:-1])}, and {unique[-1]}"

    previous_score = to_score(previous.get("risk_score"))
    current_score = to_score(snapshot.get("risk_score"))

    active_changes = []
    eased_changes = []

    active_channels = []
    eased_channels = []

    # -----------------------------
    # Credit driver: HY OAS level vs momentum
    # -----------------------------
    prev_credit_level, prev_credit_momentum = credit_components(previous)
    curr_credit_level, curr_credit_momentum = credit_components(snapshot)

    level_turned_on = (not prev_credit_level) and curr_credit_level
    momentum_turned_on = (not prev_credit_momentum) and curr_credit_momentum

    level_turned_off = prev_credit_level and (not curr_credit_level)
    momentum_turned_off = prev_credit_momentum and (not curr_credit_momentum)

    if level_turned_on or momentum_turned_on:
        if curr_credit_level and curr_credit_momentum:
            if level_turned_on and momentum_turned_on:
                active_changes.append("credit spreads were both elevated and widening")
            elif momentum_turned_on and not level_turned_on:
                active_changes.append("credit momentum deteriorated while spreads remained elevated")
            elif level_turned_on and not momentum_turned_on:
                active_changes.append("credit spreads entered stress territory")
        elif level_turned_on:
            active_changes.append("credit spreads entered stress territory")
        elif momentum_turned_on:
            active_changes.append("credit momentum deteriorated")

        active_channels.append("credit")

    if level_turned_off or momentum_turned_off:
        if (not curr_credit_level) and (not curr_credit_momentum):
            if level_turned_off and momentum_turned_off:
                eased_changes.append(
                    "credit stress eased as spreads moved below stress territory and momentum improved"
                )
            elif level_turned_off:
                eased_changes.append("credit spreads moved below stress territory")
            elif momentum_turned_off:
                eased_changes.append("credit momentum improved")
        elif level_turned_off:
            eased_changes.append("credit spreads moved below stress territory")
        elif momentum_turned_off:
            eased_changes.append("credit momentum improved")

        eased_channels.append("credit")

    # -----------------------------
    # Financial conditions: NFCI
    # -----------------------------
    previous_nfci = to_bool(previous.get("nfci_flag"))
    current_nfci = to_bool(snapshot.get("nfci_flag"))

    if (not previous_nfci) and current_nfci:
        active_changes.append("financial conditions tightened")
        active_channels.append("financial conditions")
    elif previous_nfci and (not current_nfci):
        eased_changes.append("financial conditions eased")
        eased_channels.append("financial conditions")

    # -----------------------------
    # Volatility: VIX
    # -----------------------------
    previous_vix = to_bool(previous.get("vix_flag"))
    current_vix = to_bool(snapshot.get("vix_flag"))

    if (not previous_vix) and current_vix:
        active_changes.append("volatility stress rose")
        active_channels.append("volatility")
    elif previous_vix and (not current_vix):
        eased_changes.append("volatility stress declined")
        eased_channels.append("volatility")

    # -----------------------------
    # Labor market: Initial Claims
    # -----------------------------
    previous_claims = to_bool(previous.get("claims_flag"))
    current_claims = to_bool(snapshot.get("claims_flag"))

    if (not previous_claims) and current_claims:
        active_changes.append("labor-market stress increased")
        active_channels.append("labor-market")
    elif previous_claims and (not current_claims):
        eased_changes.append("labor-market stress eased")
        eased_channels.append("labor-market")

    # -----------------------------
    # Summary phrase logic for Page 1
    # -----------------------------
    if active_changes and not eased_changes:
        if len(active_changes) == 1:
            return active_changes[0]
        if len(active_changes) == 2:
            return join_two(active_changes)

        channels = format_channels(active_channels)
        return f"risk stress broadened across {channels} channels"

    if eased_changes and not active_changes:
        if len(eased_changes) == 1:
            return eased_changes[0]
        if len(eased_changes) == 2:
            return join_two(eased_changes)

        channels = format_channels(eased_channels)
        return f"multiple stress signals eased across {channels} channels"

    if active_changes and eased_changes:
        if (
            previous_score is not None
            and current_score is not None
            and current_score > previous_score
        ):
            return f"net stress increased, led by {active_changes[0]}"

        if (
            previous_score is not None
            and current_score is not None
            and current_score < previous_score
        ):
            return f"net stress eased, led by {eased_changes[0]}"

        return "underlying stress signals were mixed"

    # Fallback if flags did not identify a driver.
    if (
        previous_score is not None
        and current_score is not None
        and current_score > previous_score
    ):
        return "risk conditions deteriorated"

    if (
        previous_score is not None
        and current_score is not None
        and current_score < previous_score
    ):
        return "risk conditions eased"

    return "risk conditions were broadly unchanged"


def build_rates_inflation_change_driver(snapshot: dict, previous: dict) -> str:
    """
    Create a Page 1 What Changed phrase for the Rates & Inflation regime.

    Uses:
        - previous/current Rates & Inflation regime
        - Rate Pressure ON/OFF
        - Inflation Concern ON/OFF

    Output is designed for the What Changed section:
        "Remained Rate Pressure Only."
        "Shifted from Rate Pressure Only to Rate + Inflation Pressure as Inflation Concern turned ON."
    """

    previous_display = format_rates_inflation_display(previous.get("rates_regime"))
    current_display = snapshot.get("rates_inflation_display", "N/A")

    previous_rate = previous.get("rate_pressure", "N/A")
    current_rate = snapshot.get("rate_pressure", "N/A")

    previous_inflation = previous.get("inflation_concern", "N/A")
    current_inflation = snapshot.get("inflation_concern", "N/A")

    if current_display == "N/A":
        return "Regime availability should be reviewed."

    if previous_display == "N/A":
        return f"Current backdrop is {current_display}."

    rate_changed = (
        previous_rate not in ["", "N/A", None]
        and current_rate not in ["", "N/A", None]
        and previous_rate != current_rate
    )

    inflation_changed = (
        previous_inflation not in ["", "N/A", None]
        and current_inflation not in ["", "N/A", None]
        and previous_inflation != current_inflation
    )

    change_phrases = []

    if rate_changed:
        if current_rate == "ON":
            change_phrases.append("Rate Pressure turned ON")
        elif current_rate == "OFF":
            change_phrases.append("Rate Pressure turned OFF")
        else:
            change_phrases.append("Rate Pressure changed")

    if inflation_changed:
        if current_inflation == "ON":
            change_phrases.append("Inflation Concern turned ON")
        elif current_inflation == "OFF":
            change_phrases.append("Inflation Concern turned OFF")
        else:
            change_phrases.append("Inflation Concern changed")

    if previous_display != current_display:
        if change_phrases:
            if len(change_phrases) == 1:
                driver_text = change_phrases[0]
            else:
                driver_text = " and ".join(change_phrases)

            return (
                f"Shifted from {previous_display} to {current_display} "
                f"as {driver_text}."
            )

        return f"Shifted from {previous_display} to {current_display}."

    if rate_changed or inflation_changed:
        if len(change_phrases) == 1:
            driver_text = change_phrases[0]
        else:
            driver_text = " and ".join(change_phrases)

        return f"Remained {current_display}, but {driver_text}."

    return f"Remained {current_display}."


def build_page1_display_fields(snapshot: dict, previous: dict, triggers: dict) -> list[list[str]]:
    """
    Build atomic Page 1 display fields for PDF_MEMO_1.

    Output structure:
        Q = Field ID
        R = Label
        S = Explanation

    Excel template controls all layout and styling.
    Python only generates data-driven label/explanation content.
    """
    risk_driver = build_risk_appetite_driver(snapshot, previous)

    previous_risk = previous.get("risk_regime", "")
    current_risk = snapshot.get("risk_regime", "N/A")

    previous_score = previous.get("risk_score", "")
    current_score = snapshot.get("risk_score", "")

    current_rates_display = snapshot["rates_inflation_display"]

    def score_display(value) -> str:
        if value in ["", None, "N/A"]:
            return "N/A"
        return f"{value} of 4"

    def score_number(value):
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return None

    def stance_tuple(source: dict) -> tuple:
        return (
            source.get("equity_stance", "N/A"),
            source.get("duration_stance", "N/A"),
            source.get("cash_stance", "N/A"),
            source.get("tips_stance", "N/A"),
        )

    current_stance = stance_tuple(snapshot)
    previous_stance = stance_tuple(previous)

    previous_score_num = score_number(previous_score)
    current_score_num = score_number(current_score)

    current_stance_text = (
        f"Equity {snapshot['equity_stance']}, "
        f"Duration {snapshot['duration_stance']}, "
        f"Cash {snapshot['cash_stance']}, "
        f"TIPS {snapshot['tips_stance']}"
    )

    # -----------------------------
    # Executive Summary
    # -----------------------------
    summary_portfolio = f"{current_stance_text}; cash remains the primary buffer."

    if previous_risk and previous_risk != "N/A" and previous_risk != current_risk:
        summary_risk = (
            f"{current_risk} — shifted from {previous_risk} as {risk_driver}."
        )
    elif current_risk == "Risk-On":
        summary_risk = "Risk-On — broad market-stress signals are not active."
    elif current_risk == "Risk-Off":
        summary_risk = "Risk-Off — broad market-stress signals are active."
    else:
        summary_risk = (
            f"{current_risk} — risk conditions are no longer cleanly Risk-On."
        )

    if snapshot["inflation_concern"] == "ON":
        inflation_text = "inflation concern is elevated"
    elif snapshot["inflation_concern"] == "OFF":
        inflation_text = "inflation concern is not elevated"
    else:
        inflation_text = "inflation concern requires review"

    summary_rates = f"{current_rates_display} — {inflation_text}."

    if snapshot["real_yield_tightening"] == "ON":
        summary_real_yield = (
            "Tightening ON — aggressive equity and duration overweights remain constrained."
        )
    elif snapshot["real_yield_tightening"] == "OFF":
        summary_real_yield = (
            "Tightening OFF — the real-yield constraint is not currently active."
        )
    else:
        summary_real_yield = "N/A — portfolio constraints require data review."

    summary_data_quality = f"{snapshot['data_quality']}."

    # -----------------------------
    # Decision Rationale
    # -----------------------------
    if current_risk == "Risk-On":
        rationale_risk = (
            "Constructive risk appetite supports equity exposure, but not an aggressive overweight."
        )
    elif current_risk == "Risk-Off":
        rationale_risk = (
            "Defensive risk appetite argues for reducing equity risk and preserving liquidity."
        )
    else:
        rationale_risk = (
            "Risk appetite deterioration argues against adding aggressive equity risk."
        )

    if snapshot["rate_pressure"] == "ON":
        if snapshot["inflation_concern"] == "ON":
            rationale_rates = (
                "Active rate pressure keeps duration neutral; inflation concern is elevated."
            )
        elif snapshot["inflation_concern"] == "OFF":
            rationale_rates = (
                "Active rate pressure keeps duration neutral; inflation concern is not elevated."
            )
        else:
            rationale_rates = (
                "Active rate pressure keeps duration neutral; inflation status requires review."
            )
    else:
        if snapshot["inflation_concern"] == "ON":
            rationale_rates = (
                "Inflation concern is elevated, but rate pressure is not active."
            )
        elif snapshot["inflation_concern"] == "OFF":
            rationale_rates = (
                "Rate pressure and inflation concern are both inactive."
            )
        else:
            rationale_rates = (
                "Rates and inflation signals require review."
            )

    if snapshot["real_yield_tightening"] == "ON":
        rationale_real_yield = (
            "Real Yield Tightening caps aggressive equity and duration overweight positions."
        )
    elif snapshot["real_yield_tightening"] == "OFF":
        rationale_real_yield = (
            "With Real Yield Tightening OFF, the real-yield constraint is less restrictive."
        )
    else:
        rationale_real_yield = (
            "Real Yield Tightening status requires review before interpreting portfolio constraints."
        )

    if snapshot["cash_stance"] == "Overweight":
        rationale_cash = (
            "Cash remains overweight as a liquidity buffer while rate pressure and real-yield tightening remain active."
        )
    else:
        rationale_cash = (
            f"Cash is {snapshot['cash_stance']}, consistent with the current risk and rates backdrop."
        )

    # -----------------------------
    # What Changed
    # -----------------------------
    if previous_risk and previous_risk != "N/A" and previous_risk != current_risk:
        change_risk = (
            f"Moved from {previous_risk} to {current_risk} as {risk_driver}."
        )
    else:
        change_risk = f"Remained {current_risk}."

    if (
        previous_score
        and previous_score != "N/A"
        and current_score
        and current_score != "N/A"
        and previous_score != current_score
    ):
        if (
            previous_score_num is not None
            and current_score_num is not None
            and current_score_num > previous_score_num
        ):
            change_score = (
                f"Increased from {score_display(previous_score)} to {score_display(current_score)}."
            )
        elif (
            previous_score_num is not None
            and current_score_num is not None
            and current_score_num < previous_score_num
        ):
            change_score = (
                f"Decreased from {score_display(previous_score)} to {score_display(current_score)}."
            )
        else:
            change_score = (
                f"Moved from {score_display(previous_score)} to {score_display(current_score)}."
            )
    else:
        change_score = f"Remained {score_display(current_score)}."

    # Always use the Rates & Inflation change helper.
    # It handles both changed and unchanged regimes.
    change_rates = build_rates_inflation_change_driver(snapshot, previous)

    if (
        previous_stance != ("N/A", "N/A", "N/A", "N/A")
        and previous_stance != current_stance
    ):
        change_portfolio = f"Shifted to {current_stance_text}."
    else:
        change_portfolio = f"Remained unchanged: {current_stance_text}."

    # -----------------------------
    # Next Triggers
    # -----------------------------
    trigger_risk = triggers["risk_appetite_trigger"]

    if snapshot["inflation_concern"] == "OFF":
        trigger_rates = (
            "Inflation concern turning ON would shift the backdrop to Rate + Inflation Pressure."
        )
    elif snapshot["inflation_concern"] == "ON":
        trigger_rates = (
            "Easing inflation concern would reduce pressure in the rates/inflation backdrop."
        )
    else:
        trigger_rates = (
            "Rates and inflation data should be reviewed before interpreting the next trigger."
        )

    trigger_real_yield = triggers["real_yield_trigger"].replace("Real Yield: ", "")

    if current_risk == "Risk-Off":
        trigger_portfolio = (
            "Easing risk stress would allow the framework to move away from a defensive stance."
        )
    else:
        trigger_portfolio = (
            "Further risk deterioration would justify a more defensive allocation stance."
        )

    # -----------------------------
    # Rule Map / Footer Note
    # -----------------------------
    rule_map = (
        "Risk Appetite Score: 0 = Risk-On | 1–2 = Neutral | 3–4 = Risk-Off. "
        "Rates & Inflation: Rate Pressure × Inflation Concern."
    )

    footer_note = (
        "Signals are regime indicators, not return forecasts. Inputs include HY OAS, NFCI, VIX, "
        "Initial Claims, 10Y yield, breakeven inflation, and real yield"
    )

    return [
        ["Field", "Label", "Explanation"],

        ["Summary Portfolio", "Portfolio", summary_portfolio],
        ["Summary Risk Appetite", "Risk Appetite", summary_risk],
        ["Summary Rates & Inflation", "Rates & Inflation", summary_rates],
        ["Summary Real Yield", "Real Yield", summary_real_yield],
        ["Summary Data Quality", "Data Quality", summary_data_quality],

        ["", "", ""],

        ["Rationale Risk Appetite", "Risk Appetite", rationale_risk],
        ["Rationale Rates & Inflation", "Rates & Inflation", rationale_rates],
        ["Rationale Real Yield", "Real Yield", rationale_real_yield],
        ["Rationale Cash", "Cash", rationale_cash],

        ["", "", ""],

        ["Change Risk Appetite", "Risk Appetite", change_risk],
        ["Change Score", "Risk Appetite Score", change_score],
        ["Change Rates & Inflation", "Rates & Inflation", change_rates],
        ["Change Portfolio", "Portfolio", change_portfolio],

        ["", "", ""],

        ["Trigger Risk Appetite", "Risk Appetite", trigger_risk],
        ["Trigger Rates & Inflation", "Rates & Inflation", trigger_rates],
        ["Trigger Real Yield", "Real Yield", trigger_real_yield],
        ["Trigger Portfolio", "Portfolio", trigger_portfolio],

        ["", "", ""],

        ["Rule Map", "Rule Map", rule_map],
        ["Footer Note", "Note", footer_note],
    ]


def build_page2_driver_detail_rows(snapshot: dict, previous: dict) -> list[list[str]]:
    """
    Build Page 2 driver detail rows for PDF_MEMO_2.

    Output structure:
        U = Section
        V = Channel
        W = Indicator
        X = Previous
        Y = Current
        Z = Interpretation

    Compact notation:
        L+T = Level + Trend
    """

    def to_bool(value) -> bool:
        if value is None:
            return False

        if isinstance(value, bool):
            return value

        if isinstance(value, (int, float)):
            return value != 0

        if isinstance(value, str):
            cleaned = value.strip().lower()
            if cleaned in {"1", "true", "yes", "y", "on"}:
                return True
            if cleaned in {"0", "false", "no", "n", "off", "", "none", "n/a", "na"}:
                return False

        return False

    def on_off(value) -> str:
        return "ON" if to_bool(value) else "OFF"

    def normalize_reason(value) -> str:
        if value is None:
            return "None"

        cleaned = str(value).strip()

        if cleaned == "" or cleaned.lower() in {"none", "n/a", "na", "null"}:
            return "None"

        lowered = cleaned.lower()

        if "both" in lowered:
            return "Both"
        if "momentum" in lowered or "trend" in lowered or "4w" in lowered or "d4w" in lowered:
            return "Trend"
        if "level" in lowered or "p80" in lowered:
            return "Level"

        return cleaned

    def reason_suffix(reason: str) -> str:
        normalized = normalize_reason(reason)

        if normalized == "Both":
            return "L+T"
        if normalized == "Level":
            return "Level"
        if normalized == "Trend":
            return "Trend"
        if normalized == "None":
            return "None"

        return normalized

    def credit_display(source: dict) -> str:
        if not to_bool(source.get("credit_flag")):
            return "OFF"

        reason = reason_suffix(source.get("hy_reason"))

        if reason == "None":
            return "ON"

        return f"ON ({reason})"

    def credit_interpretation(previous_source: dict, current_source: dict) -> str:
        previous_on = to_bool(previous_source.get("credit_flag"))
        current_on = to_bool(current_source.get("credit_flag"))
        current_reason = reason_suffix(current_source.get("hy_reason"))

        if not previous_on and current_on:
            if current_reason == "L+T":
                return "Credit stress turned active; spreads were elevated and widening."
            if current_reason == "Level":
                return "Credit stress turned active; spread level breached threshold."
            if current_reason == "Trend":
                return "Credit stress turned active; credit momentum deteriorated."
            return "Credit stress turned active."

        if previous_on and not current_on:
            return "Credit stress eased."

        if current_on:
            if current_reason == "L+T":
                return "Credit stress remains active; level and trend confirmed."
            if current_reason == "Level":
                return "Credit stress remains active; level confirmed."
            if current_reason == "Trend":
                return "Credit stress remains active; trend confirmed."
            return "Credit stress remains active."

        return "Credit stress inactive."

    def binary_interpretation(
        previous_value,
        current_value,
        turned_on_text: str,
        remains_on_text: str,
        turned_off_text: str,
        remains_off_text: str,
    ) -> str:
        previous_on = to_bool(previous_value)
        current_on = to_bool(current_value)

        if not previous_on and current_on:
            return turned_on_text
        if previous_on and current_on:
            return remains_on_text
        if previous_on and not current_on:
            return turned_off_text
        return remains_off_text

    def macro_status_display(status_value, reason_value=None) -> str:
        status = status_value if status_value in {"ON", "OFF"} else on_off(status_value)
        reason = reason_suffix(reason_value)

        if status == "ON" and reason != "None":
            return f"{status} ({reason})"

        return status

    def macro_reason_interpretation(
        current_status: str,
        current_reason,
        active_text: str,
        inactive_text: str,
    ) -> str:
        reason = reason_suffix(current_reason)

        if current_status == "ON":
            if reason == "L+T":
                return f"{active_text}; level and trend confirmed."
            if reason == "Level":
                return f"{active_text}; level confirmed."
            if reason == "Trend":
                return f"{active_text}; trend confirmed."
            return f"{active_text}."

        return inactive_text

    rows = [
        ["Section", "Channel", "Indicator", "Previous", "Current", "Interpretation"],

        [
            "Risk Appetite",
            "Credit",
            "HY OAS",
            credit_display(previous),
            credit_display(snapshot),
            credit_interpretation(previous, snapshot),
        ],
        [
            "Risk Appetite",
            "Financial Conditions",
            "NFCI",
            on_off(previous.get("nfci_flag")),
            on_off(snapshot.get("nfci_flag")),
            binary_interpretation(
                previous.get("nfci_flag"),
                snapshot.get("nfci_flag"),
                "Financial conditions tightened.",
                "Financial conditions remain stressed.",
                "Financial conditions eased.",
                "Financial conditions not stressed.",
            ),
        ],
        [
            "Risk Appetite",
            "Volatility",
            "VIX",
            on_off(previous.get("vix_flag")),
            on_off(snapshot.get("vix_flag")),
            binary_interpretation(
                previous.get("vix_flag"),
                snapshot.get("vix_flag"),
                "Volatility stress rose.",
                "Volatility stress remains active.",
                "Volatility stress declined.",
                "Volatility stress inactive.",
            ),
        ],
        [
            "Risk Appetite",
            "Labor Market",
            "Initial Claims",
            on_off(previous.get("claims_flag")),
            on_off(snapshot.get("claims_flag")),
            binary_interpretation(
                previous.get("claims_flag"),
                snapshot.get("claims_flag"),
                "Labor-market stress increased.",
                "Labor-market stress remains active.",
                "Labor-market stress eased.",
                "Labor-market stress inactive.",
            ),
        ],

        [
            "Rates & Inflation",
            "Rates",
            "Rate Pressure",
            macro_status_display(previous.get("rate_pressure"), previous.get("rate_reason")),
            macro_status_display(snapshot.get("rate_pressure"), snapshot.get("rate_reason")),
            macro_reason_interpretation(
                snapshot.get("rate_pressure"),
                snapshot.get("rate_reason"),
                "Rate pressure remains active",
                "Rate pressure is inactive.",
            ),
        ],
        [
            "Rates & Inflation",
            "Inflation",
            "Inflation Concern",
            macro_status_display(previous.get("inflation_concern"), previous.get("inflation_reason")),
            macro_status_display(snapshot.get("inflation_concern"), snapshot.get("inflation_reason")),
            macro_reason_interpretation(
                snapshot.get("inflation_concern"),
                snapshot.get("inflation_reason"),
                "Inflation concern remains active",
                "Inflation concern remains inactive.",
            ),
        ],
        [
            "Rates & Inflation",
            "Real Yield",
            "Real Yield Tightening",
            macro_status_display(previous.get("real_yield_tightening"), previous.get("real_yield_reason")),
            macro_status_display(snapshot.get("real_yield_tightening"), snapshot.get("real_yield_reason")),
            macro_reason_interpretation(
                snapshot.get("real_yield_tightening"),
                snapshot.get("real_yield_reason"),
                "Real-yield tightening remains active",
                "Real-yield tightening is inactive.",
            ),
        ],
    ]

    return rows


def build_page2_risk_rule_rows() -> list[list[str]]:
    """
    Build compact Risk Appetite rule rows for PDF_MEMO_2.

    Output structure:
        U = Indicator
        V = Channel
        W = Trigger Rule
        X = Score Effect
    """
    return [
        ["Indicator", "Channel", "Trigger Rule", "Score Effect"],
        [
            "HY OAS",
            "Credit",
            "Level > rolling P80 OR positive 4W change",
            "+1 if either condition is active",
        ],
        [
            "NFCI",
            "Financial Conditions",
            "Level > rolling P80",
            "+1 when active",
        ],
        [
            "VIX",
            "Volatility",
            "Level > rolling P80",
            "+1 when active",
        ],
        [
            "Initial Claims",
            "Labor Market",
            "Level > rolling P80",
            "+1 when active",
        ],
        [
            "Risk Appetite Score",
            "Regime Mapping",
            "0 = Risk-On | 1–2 = Neutral | 3–4 = Risk-Off",
            "Maximum score = 4",
        ],
    ]


def build_page2_rates_rule_rows() -> list[list[str]]:
    """
    Build compact Rates & Inflation signal rules and regime matrix.

    Output structure:
        AA = Component
        AB = Rate Pressure
        AC = Inflation Concern
        AD = Regime
        AE = Rule / Display Label
    """
    return [
        ["Component", "Rate Pressure", "Inflation Concern", "Regime", "Rule / Display Label"],
        [
            "10Y Treasury Yield",
            "ON if level > P80 or 4W change > 0",
            "",
            "",
            "Defines Rate Pressure",
        ],
        [
            "10Y Breakeven Inflation",
            "",
            "ON if level > P80 or 4W change > 0",
            "",
            "Defines Inflation Concern",
        ],
        [
            "10Y Real Yield",
            "",
            "",
            "Constraint",
            "Real Yield Tightening is ON if level > P80 or 4W change > 0",
        ],
        [
            "Regime Matrix",
            "ON",
            "ON",
            "Regime 1",
            "Rate + Inflation Pressure",
        ],
        [
            "Regime Matrix",
            "ON",
            "OFF",
            "Regime 2",
            "Rate Pressure Only",
        ],
        [
            "Regime Matrix",
            "OFF",
            "ON",
            "Regime 3",
            "Inflation Concern Only",
        ],
        [
            "Regime Matrix",
            "OFF",
            "OFF",
            "Regime 4",
            "No Rate / Inflation Pressure",
        ],
    ]


def build_page2_risk_chart_rows(
    workbook,
    lookback_weeks: int = 52,
) -> list[list]:
    """
    Build fixed-length historical data for the Page 2 Risk Appetite chart.

    Output block:
        AG = Date
        AH = Risk Appetite Score
        AI = Neutral Boundary
        AJ = Risk-Off Boundary
        AK = Latest Score

    The function always returns:
        1 header row + lookback_weeks data rows

    Blank rows are appended when fewer than lookback_weeks observations exist,
    allowing the Excel chart to reference one permanent fixed range.
    """
    risk_rows = read_rows_by_date(workbook, "risk_appetite")
    selected_rows = risk_rows[-lookback_weeks:]

    output = [
        [
            "Date",
            "Risk Appetite Score",
            "Neutral Boundary",
            "Risk-Off Boundary",
            "Latest Score",
        ]
    ]

    for index, row in enumerate(selected_rows):
        raw_date = row.get("__date__")
        raw_score = row_value(
            row,
            ["Risk Appetite Score", "Risk_Appetite_Score"],
            None,
        )

        try:
            score = float(raw_score)
        except (TypeError, ValueError):
            score = None

        latest_score = score if index == len(selected_rows) - 1 else None

        output.append(
            [
                raw_date,
                score,
                1,
                3,
                latest_score,
            ]
        )

    while len(output) < lookback_weeks + 1:
        output.append([None, None, None, None, None])

    return output


def build_page2_rates_chart_rows(
    workbook,
    lookback_weeks: int = 52,
) -> list[list]:
    """
    Build fixed-length historical data for the Page 2 Rates & Inflation chart.

    Output block:
        AM = Date
        AN = Regime Number
        AO = Rate Pressure
        AP = Inflation Concern
        AQ = Real Yield Tightening
        AR = Latest Regime

    Binary values:
        ON  = 1
        OFF = 0

    The fixed output range allows the template chart to update automatically
    whenever run_weekly.py refreshes the workbook.
    """
    rates_rows = read_rows_by_date(workbook, "rates_inflation")
    selected_rows = rates_rows[-lookback_weeks:]

    def binary_value(value):
        formatted = format_on_off(value)

        if formatted == "ON":
            return 1
        if formatted == "OFF":
            return 0

        return None

    output = [
        [
            "Date",
            "Regime Number",
            "Rate Pressure",
            "Inflation Concern",
            "Real Yield Tightening",
            "Latest Regime",
        ]
    ]

    for index, row in enumerate(selected_rows):
        raw_date = row.get("__date__")

        raw_regime = row_value(
            row,
            ["Rates & Inflation Regime", "Rates_Inflation_Regime"],
            None,
        )
        regime_number = parse_regime_number(raw_regime)

        rate_pressure = row_value(
            row,
            ["Rate_Pressure", "Rate Pressure"],
            None,
        )
        inflation_concern = row_value(
            row,
            ["Inflation_Concern", "Inflation Concern"],
            None,
        )
        real_yield_tightening = row_value(
            row,
            ["Real_Yield_Tightening", "Real Yield Tightening"],
            None,
        )

        # Fall back to regime mapping if component fields are unavailable.
        if format_on_off(rate_pressure) == "N/A" or format_on_off(inflation_concern) == "N/A":
            inferred_rate, inferred_inflation = infer_pressure_flags(raw_regime)

            if format_on_off(rate_pressure) == "N/A":
                rate_pressure = inferred_rate

            if format_on_off(inflation_concern) == "N/A":
                inflation_concern = inferred_inflation

        latest_regime = (
            regime_number
            if index == len(selected_rows) - 1
            else None
        )

        output.append(
            [
                raw_date,
                regime_number,
                binary_value(rate_pressure),
                binary_value(inflation_concern),
                binary_value(real_yield_tightening),
                latest_regime,
            ]
        )

    while len(output) < lookback_weeks + 1:
        output.append([None, None, None, None, None, None])

    return output


def read_stress_event_rows(workbook) -> list[dict]:
    worksheet = find_sheet(workbook, SHEET_CANDIDATES["stress_event_audit"])
    if worksheet is None:
        return []

    headers = []

    for cell in worksheet[1]:
        headers.append(str(cell.value).strip() if cell.value is not None else "")

    rows = []

    for row_number in range(2, worksheet.max_row + 1):
        row_data = {}
        has_value = False

        for col_number, header in enumerate(headers, start=1):
            if not header:
                continue

            value = worksheet.cell(row=row_number, column=col_number).value

            if value is not None:
                has_value = True

            row_data[header] = value

        if has_value:
            rows.append(row_data)

    return rows


def get_row_alias(row: dict, aliases: list[str], default="N/A"):
    for alias in aliases:
        if alias in row:
            value = row[alias]
            return clean_public_text(value) if value is not None else default

    return default


def select_stress_event_rows(workbook) -> list[list[str]]:
    source_rows = read_stress_event_rows(workbook)

    priorities = [
        ("Global Financial Crisis", ["global financial", "gfc"]),
        ("COVID Shock", ["covid"]),
        ("2022 Inflation / Rates Shock", ["2022", "inflation", "rates shock"]),
        ("2023 Regional Bank Stress", ["2023", "regional bank", "bank stress"]),
    ]

    selected = []
    used_indices = set()

    for display_name, keywords in priorities:
        matched_index = None
        matched_row = None

        for index, row in enumerate(source_rows):
            if index in used_indices:
                continue

            event_text = get_row_alias(row, ["Event"], "").lower()

            if any(keyword in event_text for keyword in keywords):
                matched_index = index
                matched_row = row
                break

        if matched_row is None:
            selected.append([display_name, "N/A", "N/A", "N/A", "N/A", "N/A", "N/A"])
            continue

        used_indices.add(matched_index)

        selected.append(
            [
                clean_public_text(get_row_alias(matched_row, ["Event"], display_name)) or display_name,
                get_row_alias(matched_row, ["Diagnostic Result", "Miss_Status"]),
                get_row_alias(matched_row, ["Response Lag (Weeks)", "Response_Lag_Weeks"]),
                get_row_alias(matched_row, ["Peak Risk Appetite Regime", "Peak_Risk_Appetite_Regime"]),
                get_row_alias(matched_row, ["Peak Rates & Inflation Regime", "Peak_Rates_Inflation_Regime"]),
                get_row_alias(matched_row, ["Peak Drivers", "Peak_Drivers"]),
                get_row_alias(
                    matched_row,
                    [
                        "Pre-Existing Stress",
                        "Pre-Existing Stress at Window Start",
                        "Already_Stressed_At_Window_Start",
                    ],
                ),
            ]
        )

    return selected


def build_page2_stress_audit_rows(workbook) -> list[list]:
    """
    Build compact Historical Stress Audit rows for PDF_MEMO_2.

    Output:
        Event
        Detection
        Lag (Weeks)
        Peak Risk Score
        Peak R&I
    """

    source_rows = read_stress_event_rows(workbook)

    priorities = [
        ("Global Financial Crisis", ["global financial", "gfc"]),
        ("COVID Shock", ["covid"]),
        ("2022 Inflation / Rates Shock", ["2022", "inflation", "rates shock"]),
        ("2023 Regional Bank Stress", ["2023", "regional bank", "bank stress"]),
    ]

    def raw_alias(row: dict, aliases: list[str], default=None):
        for alias in aliases:
            if alias in row:
                value = row[alias]

                if value is not None and str(value).strip() != "":
                    return value

        return default

    def format_detection(value) -> str:
        if value is None:
            return "N/A"

        text = str(value).strip().lower()

        if text in {
            "no miss",
            "no_miss",
            "detected",
            "pass",
            "passed",
            "true",
            "1",
        }:
            return "Detected"

        if text in {
            "miss",
            "missed",
            "not detected",
            "not_detected",
            "fail",
            "failed",
            "false",
            "0",
        }:
            return "Missed"

        if "no miss" in text:
            return "Detected"

        if "not detected" in text:
            return "Missed"

        if "miss" in text:
            return "Missed"

        if "detected" in text or "pass" in text:
            return "Detected"

        return clean_public_text(value) or "N/A"

    def format_lag(value):
        if value is None:
            return "N/A"

        try:
            number = float(value)

            if number.is_integer():
                return int(number)

            return round(number, 1)

        except (TypeError, ValueError):
            cleaned = clean_public_text(value)
            return cleaned if cleaned else "N/A"

    def format_peak_score(value):
        """
        Return only a numeric 0–4 score.

        Does not convert Risk-On / Neutral / Risk-Off into a score,
        because a regime label does not reveal the exact score.
        """
        if value is None:
            return "N/A"

        try:
            number = float(value)

            if 0 <= number <= 4:
                if number.is_integer():
                    return int(number)

                return round(number, 1)

            return "N/A"

        except (TypeError, ValueError):
            text = str(value).strip()

            match = re.search(
                r"(?<!\d)([0-4](?:\.\d+)?)(?!\d)",
                text,
            )

            if match:
                number = float(match.group(1))

                if number.is_integer():
                    return int(number)

                return round(number, 1)

            return "N/A"

    def format_peak_rates_regime(value) -> str:
        regime_number = parse_regime_number(value)

        if regime_number in {1, 2, 3, 4}:
            return f"R{regime_number}"

        return "N/A"

    selected_rows = []
    used_indices = set()

    for display_name, keywords in priorities:
        matched_index = None
        matched_row = None

        for index, row in enumerate(source_rows):
            if index in used_indices:
                continue

            event_text = str(
                raw_alias(
                    row,
                    ["Event"],
                    "",
                )
            ).strip().lower()

            if any(keyword in event_text for keyword in keywords):
                matched_index = index
                matched_row = row
                break

        if matched_row is None:
            selected_rows.append(
                [
                    display_name,
                    "N/A",
                    "N/A",
                    "N/A",
                    "N/A",
                ]
            )
            continue

        used_indices.add(matched_index)

        event_name = clean_public_text(
            raw_alias(
                matched_row,
                ["Event"],
                display_name,
            )
        )

        diagnostic_result = raw_alias(
            matched_row,
            [
                "Diagnostic Result",
                "Miss_Status",
                "Miss Status",
                "Detection Result",
                "Stress Detection",
            ],
            None,
        )

        response_lag = raw_alias(
            matched_row,
            [
                "Response Lag (Weeks)",
                "Response_Lag_Weeks",
                "Response Lag",
                "Detection Lag (Weeks)",
                "Detection_Lag_Weeks",
            ],
            None,
        )

        peak_risk_score = raw_alias(
            matched_row,
            [
                "Peak Risk Appetite Score",
                "Peak_Risk_Appetite_Score",
                "Peak Risk Score",
                "Peak_Risk_Score",
                "Peak Risk Appetite Score",
                "Peak_Risk_Appetite_Score",
                "Peak Score",
            ],
            None,
        )

        peak_rates_regime = raw_alias(
            matched_row,
            [
                "Peak Rates & Inflation Regime",
                "Peak_Rates_Inflation_Regime",
                "Peak Rates & Inflation Regime",
                "Peak_Rates_Inflation_Regime",
                "Peak Rates & Inflation",
            ],
            None,
        )

        selected_rows.append(
            [
                event_name or display_name,
                format_detection(diagnostic_result),
                format_lag(response_lag),
                format_peak_score(peak_risk_score),
                format_peak_rates_regime(peak_rates_regime),
            ]
        )

    return [
        [
            "Event",
            "Detection",
            "Lag (Weeks)",
            "Peak Risk Score",
            "Peak R&I",
        ],
        *selected_rows,
    ]


def reset_pdf_data_sheet(workbook):
    if PDF_DATA_SHEET in workbook.sheetnames:
        worksheet = workbook[PDF_DATA_SHEET]
        worksheet.delete_rows(1, worksheet.max_row)
    else:
        worksheet = workbook.create_sheet(PDF_DATA_SHEET)

    return worksheet


def write_row(worksheet, row_number: int, start_column: int, values: list):
    for offset, value in enumerate(values):
        worksheet.cell(row=row_number, column=start_column + offset).value = value


def style_pdf_data_sheet(worksheet, keep_visible: bool) -> None:
    navy_fill = PatternFill("solid", fgColor="1F4E78")
    white_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_ranges = [
        "A1:C1",
        "E1:G1",
        "I1:O1",
        "Q1:R1",
        "A18:C18",
        "A26:C26",
        "A41:C41",
        "Q7:R7",
        "Q14:S14",
        "U1:Z1",
        "U12:X12",
        "AA12:AE12",
        "AG1:AK1",
        "AM1:AR1",
        "AT1:AX1",
    ]

    for cell_range in header_ranges:
        for row in worksheet[cell_range]:
            for cell in row:
                cell.fill = navy_fill
                cell.font = white_font
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
                cell.border = border

    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(
                wrap_text=True,
                vertical="top",
            )

    widths = {
        "A": 12,
        "B": 32,
        "C": 80,
        "D": 4,

        "E": 16,
        "F": 20,
        "G": 58,
        "H": 4,

        "I": 30,
        "J": 20,
        "K": 18,
        "L": 24,
        "M": 28,
        "N": 54,
        "O": 20,
        "P": 4,

        "Q": 34,
        "R": 24,
        "S": 90,
        "T": 4,

        # Page 2 driver detail and Risk Appetite rule source.
        "U": 22,
        "V": 24,
        "W": 44,
        "X": 28,
        "Y": 20,
        "Z": 70,

        # Page 2 Rates & Inflation rule source.
        "AA": 28,
        "AB": 32,
        "AC": 32,
        "AD": 18,
        "AE": 42,

        "AF": 4,

        # Risk Appetite chart source.
        "AG": 14,
        "AH": 20,
        "AI": 18,
        "AJ": 18,
        "AK": 16,

        "AL": 4,

        # Rates & Inflation chart source.
        "AM": 14,
        "AN": 18,
        "AO": 18,
        "AP": 20,
        "AQ": 22,
        "AR": 16,

        "AS": 4,

        # Historical Stress Audit source.
        "AT": 34, 
        "AU": 18, 
        "AV": 16, 
        "AW": 20, 
        "AX": 16,
    }

    for column_letter, width in widths.items():
        worksheet.column_dimensions[column_letter].width = width

    # Date formatting for Page 2 chart-source blocks.
    for row_number in range(2, 54):
        worksheet.cell(
            row=row_number,
            column=33,
        ).number_format = "mmm d, yyyy"

        worksheet.cell(
            row=row_number,
            column=39,
        ).number_format = "mmm d, yyyy"

    # Default row height.
    for row_number in range(1, worksheet.max_row + 1):
        worksheet.row_dimensions[row_number].height = 20

    # Existing long narrative rows from the original _PDF_DATA contract.
    original_long_rows = [
        3,   # Memo Subtitle
        8,   # Evidence summary
        9,   # Evidence summary
        10,  # Evidence summary
        11,  # Evidence summary
        12,  # Evidence summary / rule headers
        15,  # Overall Decision Summary
        17,  # Executive Summary
        19,  # Executive Interpretation
        20,  # What Changed
        21,  # Next Triggers
        22,  # Key Risks and Limitations
        23,  # Footer
        38,  # Change Summary
        46,  # Trigger Summary
    ]

    for row_number in original_long_rows:
        worksheet.row_dimensions[row_number].height = 60

    # Page 1 atomic display rows in Q:S.
    # Row height applies across the entire worksheet, so this also gives
    # sufficient space to overlapping Page 2 rule-source rows.
    for row_number in range(14, 40):
        worksheet.row_dimensions[row_number].height = 30

    # Page 2 driver-detail rows.
    for row_number in range(1, 9):
        worksheet.row_dimensions[row_number].height = 28

    # Header rows.
    worksheet.row_dimensions[1].height = 22
    worksheet.row_dimensions[12].height = 22
    worksheet.row_dimensions[14].height = 22

    worksheet.freeze_panes = "A2"
    worksheet.sheet_state = "visible" if keep_visible else "hidden"


def update_pdf_data(output_path: str, keep_visible: bool = True) -> None:
    """
    Update _PDF_DATA inside the Excel output workbook.

    _PDF_DATA is the automated source layer for PDF_MEMO_1 and PDF_MEMO_2.
    Public-facing workbook text uses Risk Appetite, Rates & Inflation, and Portfolio Router terminology.
    """
    path = Path(output_path)

    if not path.exists():
        raise FileNotFoundError(f"Workbook not found: {path}")

    workbook = load_workbook(path)
    worksheet = reset_pdf_data_sheet(workbook)

    snapshot, previous = build_snapshot(workbook)
    triggers = build_triggers(snapshot)
    evidence = build_evidence_summaries(snapshot)
    page1_display_rows = build_page1_display_fields(snapshot, previous, triggers)
    page2_driver_rows = build_page2_driver_detail_rows(snapshot, previous)
    page2_risk_rule_rows = build_page2_risk_rule_rows()
    page2_rates_rule_rows = build_page2_rates_rule_rows()
    page2_risk_chart_rows = build_page2_risk_chart_rows(
        workbook,
        lookback_weeks=52,
        )

    page2_rates_chart_rows = build_page2_rates_chart_rows(
        workbook,
        lookback_weeks=52,
        )
    
    page2_stress_audit_rows = build_page2_stress_audit_rows(workbook)


    memo_title = "Macro Regime-to-Portfolio Decision Framework"
    memo_subtitle = (
        f"Weekly Research Memo | As of: {snapshot['as_of_date']} | "
        "Data: FRED | Tool: Python, Excel | Frequency: Weekly"
    )

    decision_summary = build_decision_summary(snapshot)
    executive_summary = build_executive_summary(snapshot)
    executive_interpretation = build_executive_interpretation(snapshot)
    change_summary = build_change_summary(snapshot, previous)
    key_risks = build_key_risks_and_limitations()
    footer = (
        "This memo summarizes a rule-based macro regime framework. "
        "It is intended for research and decision-support purposes only."
    )

    write_row(worksheet, 1, 1, ["Section", "Field", "Value"])

    snapshot_rows = [
        ["Header", "Memo Title", memo_title],
        ["Header", "Memo Subtitle", memo_subtitle],
        ["Header", "As of Date", snapshot["as_of_date"]],
        ["Snapshot", "Risk Appetite Regime", snapshot["risk_regime"]],
        ["Snapshot", "Risk Appetite Score", snapshot["risk_score"]],
        ["Snapshot", "Risk Appetite Trend", snapshot["risk_trend"]],
        ["Snapshot", "Rates & Inflation Regime", snapshot["rates_regime"]],
        ["Snapshot", "Rates & Inflation Regime Label", snapshot["rates_regime_label"]],
        ["Snapshot", "Rate Pressure", snapshot["rate_pressure"]],
        ["Snapshot", "Inflation Concern", snapshot["inflation_concern"]],
        ["Snapshot", "Real Yield Tightening", snapshot["real_yield_tightening"]],
        ["Snapshot", "Confidence", snapshot["confidence"]],
        ["Snapshot", "Data Quality Summary", snapshot["data_quality"]],
        ["Snapshot", "Overall Decision Summary", decision_summary],
        ["Snapshot", "Rates & Inflation Display", snapshot["rates_inflation_display"]],
        ["Snapshot", "Executive Summary", executive_summary],
    ]

    for row_index, row_values in enumerate(snapshot_rows, start=2):
        write_row(worksheet, row_index, 1, row_values)

    write_row(worksheet, 1, 5, ["Asset Class", "Stance", "Band"])

    portfolio_note = (
        "Equity and duration overweights are capped while Real Yield Tightening remains ON."
        if snapshot["real_yield_tightening"] == "ON"
        else "Real Yield Tightening is not currently capping aggressive equity or duration overweights."
    )

    portfolio_rows = [
        ["Equity", snapshot["equity_stance"], snapshot["equity_band"]],
        ["Duration", snapshot["duration_stance"], snapshot["duration_band"]],
        ["Cash", snapshot["cash_stance"], snapshot["cash_band"]],
        ["TIPS", snapshot["tips_stance"], snapshot["tips_band"]],
        ["Note", "Real Yield Constraint", portfolio_note],
    ]

    for row_index, row_values in enumerate(portfolio_rows, start=2):
        write_row(worksheet, row_index, 5, row_values)

    write_row(worksheet, 18, 1, ["Section", "Field", "Text"])

    narrative_rows = [
        ["Page 1", "Executive Interpretation", executive_interpretation],
        ["Page 1", "What Changed", change_summary],
        ["Page 1", "Next Triggers", triggers["trigger_summary"]],
        ["Page 1", "Key Risks and Limitations", key_risks],
        ["Page 1", "Footer", footer],
    ]

    for row_index, row_values in enumerate(narrative_rows, start=19):
        write_row(worksheet, row_index, 1, row_values)

    write_row(worksheet, 26, 1, ["Section", "Field", "Value"])

    change_rows = [
        ["Change", "Previous As of Date", previous.get("as_of_date", "")],
        ["Change", "Previous Risk Appetite Regime", previous.get("risk_regime", "")],
        ["Change", "Previous Risk Appetite Score", previous.get("risk_score", "")],
        ["Change", "Previous Rates & Inflation Regime", previous.get("rates_regime", "")],
        ["Change", "Previous Rate Pressure", previous.get("rate_pressure", "")],
        ["Change", "Previous Inflation Concern", previous.get("inflation_concern", "")],
        ["Change", "Previous Real Yield Tightening", previous.get("real_yield_tightening", "")],
        ["Change", "Previous Equity Stance", previous.get("equity_stance", "")],
        ["Change", "Previous Duration Stance", previous.get("duration_stance", "")],
        ["Change", "Previous Cash Stance", previous.get("cash_stance", "")],
        ["Change", "Previous TIPS Stance", previous.get("tips_stance", "")],
        ["Change", "Change Summary", change_summary],
    ]

    for row_index, row_values in enumerate(change_rows, start=27):
        write_row(worksheet, row_index, 1, row_values)

    write_row(worksheet, 41, 1, ["Section", "Field", "Value"])

    trigger_rows = [
        ["Trigger", "Risk Appetite Trigger", triggers["risk_appetite_trigger"]],
        ["Trigger", "Rates & Inflation Trigger", triggers["rates_inflation_trigger"]],
        ["Trigger", "Real Yield Trigger", triggers["real_yield_trigger"]],
        ["Trigger", "Portfolio Constraint Trigger", triggers["portfolio_constraint_trigger"]],
        ["Trigger", "Trigger Summary", triggers["trigger_summary"]],
    ]

    for row_index, row_values in enumerate(trigger_rows, start=42):
        write_row(worksheet, row_index, 1, row_values)

    write_row(
        worksheet,
        1,
        9,
        [
            "Event",
            "Diagnostic Result",
            "Response Lag",
            "Peak Risk Appetite",
            "Peak Rates & Inflation",
            "Peak Drivers",
            "Pre-Existing Stress",
        ],
    )

    for row_index, row_values in enumerate(select_stress_event_rows(workbook), start=2):
        write_row(worksheet, row_index, 9, row_values)

    write_row(worksheet, 1, 17, ["Field", "Text"])

    page_2_rows = [
        [
            "Supporting Evidence Note",
            "Charts summarize the key signal history and current macro constraints behind the portfolio stance.",
        ],
        ["Rule Rationale", build_rule_rationale()],
        ["Limitations", build_limitations()],
        [
            "Chart Selection Note",
            "Current evidence emphasizes Risk Appetite, Rates & Inflation regime history, real-yield pressure, and the 10Y Treasury yield as the primary rate-pressure indicator.",
        ],
    ]

    for row_index, row_values in enumerate(page_2_rows, start=2):
        write_row(worksheet, row_index, 17, row_values)

    write_row(worksheet, 7, 17, ["Field", "Text"])

    evidence_rows = [
        ["Risk Appetite Evidence Summary", evidence["risk_appetite_evidence"]],
        ["Rates & Inflation Evidence Summary", evidence["rates_inflation_evidence"]],
        ["Real Yield Evidence Summary", evidence["real_yield_evidence"]],
        ["Portfolio Evidence Summary", evidence["portfolio_evidence"]],
        ["Chart Evidence Summary", evidence["chart_evidence"]],
    ]

    for row_index, row_values in enumerate(evidence_rows, start=8):
        write_row(worksheet, row_index, 17, row_values)

    # Page 1 display fields for PDF_MEMO_1.
    # Q14:S37 stores atomic, layout-friendly memo lines.
    for row_index, row_values in enumerate(page1_display_rows, start=14):
        write_row(worksheet, row_index, 17, row_values)

    # Page 2 driver detail table for PDF_MEMO_2.
    # U1:Z8 stores driver/audit evidence rows.
    for row_index, row_values in enumerate(page2_driver_rows, start=1):
        write_row(worksheet, row_index, 21, row_values)

    # Page 2 compact Risk Appetite rule block.
    # U12:X17
    for row_index, row_values in enumerate(page2_risk_rule_rows, start=12):
        write_row(worksheet, row_index, 21, row_values)

    # Page 2 compact Rates & Inflation rule block.
    # AA12:AE19
    for row_index, row_values in enumerate(page2_rates_rule_rows, start=12):
        write_row(worksheet, row_index, 27, row_values)

    # Page 2 Risk Appetite historical chart source.
    # AG1:AK53 = header + latest 52 weekly observations.
    for row_index, row_values in enumerate(
        page2_risk_chart_rows,
        start=1,
    ):
        write_row(
            worksheet,
            row_index,
            33,
            row_values,
        )

    # Page 2 Rates & Inflation historical chart source.
    # AM1:AR53 = header + latest 52 weekly observations.
    for row_index, row_values in enumerate(
        page2_rates_chart_rows,
        start=1,
    ):
        write_row(
            worksheet,
            row_index,
            39,
            row_values,
        )
    
    # Page 2 compact Historical Stress Audit source.
    # AT1:AX5 = header + four stress events.
    for row_index, row_values in enumerate(
        page2_stress_audit_rows,
        start=1,
    ):
        write_row(
            worksheet,
            row_index,
            46,
            row_values,
        )


    style_pdf_data_sheet(worksheet, keep_visible=keep_visible)
    workbook.save(path)


def sync_pdf_data_to_template(
    source_output_path: str = "outputs/Macro_Regime_Framework_Output.xlsx",
    template_path: str = "templates/dashboard_template.xlsx",
) -> None:
    """
    Copy the generated _PDF_DATA sheet from the latest output workbook
    into the dashboard template workbook.

    Purpose:
    - This is only for design preview.
    - It lets PDF_MEMO_1 and PDF_MEMO_2 display realistic text while editing
      templates/dashboard_template.xlsx.
    - The actual production data still comes from update_pdf_data() during run_weekly.py.
    """
    source_path = Path(source_output_path)
    target_path = Path(template_path)

    if not source_path.exists():
        raise FileNotFoundError(f"Source output workbook not found: {source_path}")

    if not target_path.exists():
        raise FileNotFoundError(f"Template workbook not found: {target_path}")

    source_wb = load_workbook(source_path)
    target_wb = load_workbook(target_path)

    if PDF_DATA_SHEET not in source_wb.sheetnames:
        raise ValueError(f"{PDF_DATA_SHEET} not found in source workbook: {source_path}")

    source_ws = source_wb[PDF_DATA_SHEET]

    if PDF_DATA_SHEET in target_wb.sheetnames:
        target_ws = target_wb[PDF_DATA_SHEET]
        target_ws.delete_rows(1, target_ws.max_row)
    else:
        target_ws = target_wb.create_sheet(PDF_DATA_SHEET)

    for row in source_ws.iter_rows():
        for source_cell in row:
            target_cell = target_ws.cell(
                row=source_cell.row,
                column=source_cell.column,
                value=source_cell.value,
            )

            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.fill = copy(source_cell.fill)
                target_cell.border = copy(source_cell.border)
                target_cell.alignment = copy(source_cell.alignment)
                target_cell.number_format = source_cell.number_format
                target_cell.protection = copy(source_cell.protection)

    for column_letter, dimension in source_ws.column_dimensions.items():
        target_ws.column_dimensions[column_letter].width = dimension.width
        target_ws.column_dimensions[column_letter].hidden = dimension.hidden

    for row_number, dimension in source_ws.row_dimensions.items():
        target_ws.row_dimensions[row_number].height = dimension.height
        target_ws.row_dimensions[row_number].hidden = dimension.hidden

    target_ws.freeze_panes = source_ws.freeze_panes
    target_ws.sheet_state = "visible"

    target_wb.save(target_path)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        raise SystemExit(
            "Usage:\n"
            "  python pdf_memo.py <path-to-output-workbook>\n"
            "  python pdf_memo.py --sync-template\n"
            "  python pdf_memo.py --sync-template <source-output-workbook> <template-workbook>"
        )

    if sys.argv[1] == "--sync-template":
        source = (
            sys.argv[2]
            if len(sys.argv) >= 3
            else "outputs/Macro_Regime_Framework_Output.xlsx"
        )
        template = (
            sys.argv[3]
            if len(sys.argv) >= 4
            else "templates/dashboard_template.xlsx"
        )
        sync_pdf_data_to_template(source, template)
    else:
        update_pdf_data(sys.argv[1], keep_visible=True)